#!/usr/bin/env python3
"""Mapa Mestre do Sistema Financeiro - Parte 1: Setup + Abas 1-6"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

WB_PATH = os.path.join(os.path.dirname(__file__), "Mapa_Mestre_Sistema_Financeiro.xlsx")

# === STYLES ===
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)

def auto_width(ws, num_cols, max_w=45):
    for col in range(1, num_cols + 1):
        mx = 10
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(col)].width = mx + 2

def add_sheet(wb, title, headers, data):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    style_header(ws, len(headers))
    for row_data in data:
        ws.append(row_data)
    if data:
        style_data(ws, 2, len(data) + 1, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =====================================================
# ABA 0: GUIA DE LEITURA
# =====================================================
headers = ["Topico", "Explicacao"]
data = [
    ["O que e este arquivo?", "Este Excel e o Mapa Mestre do sistema financeiro. Ele documenta TODAS as tabelas, campos, regras, fluxos, telas e integracoes do software."],
    ["Para quem e?", "Para gestores, controllers, analistas e qualquer pessoa que precise entender como o sistema funciona sem precisar ler codigo."],
    ["Por onde comecar?", "Comece pela aba 'Visao Geral' para entender os modulos. Depois va para 'Glossario' se nao entender algum termo tecnico."],
    ["O que e uma TABELA?", "E onde os dados ficam guardados no banco de dados. Ex: a tabela 'Supplier' guarda todos os fornecedores cadastrados."],
    ["O que e um CAMPO?", "E uma coluna dentro de uma tabela. Ex: na tabela Supplier, o campo 'name' guarda o nome do fornecedor."],
    ["O que e uma REGRA DE NEGOCIO?", "E uma condicao ou logica que o sistema aplica automaticamente. Ex: 'Um lancamento so pode ser incorporado se estiver validado'."],
    ["O que e um FLUXO?", "E o caminho que a informacao percorre: de onde ela entra, por onde passa, e onde e armazenada."],
    ["O que e CHAVE PRIMARIA (PK)?", "E o identificador unico de cada registro. Como o CPF de uma pessoa - nao se repete. Toda tabela tem uma."],
    ["O que e CHAVE ESTRANGEIRA (FK)?", "E um campo que conecta uma tabela a outra. Ex: o campo 'tenantId' no Supplier aponta para qual empresa ele pertence."],
    ["O que e 1:N?", "Relacionamento 'um para muitos'. Ex: 1 Empresa tem N Fornecedores. Uma empresa pode ter varios fornecedores."],
    ["O que e ENUM?", "Lista fixa de opcoes. Ex: Status pode ser OPEN, PARTIAL, SETTLED, CANCELLED - so essas opcoes sao validas."],
    ["Como usar para conversar com devs?", "Use as abas 'Dicionario de Dados' e 'Tabelas' para falar os nomes corretos. Use 'Regras de Negocio' para explicar o que espera."],
    ["Como usar para pedir melhorias?", "Consulte 'Backlog de Melhorias' e 'Problemas/Riscos'. Referencie IDs das regras e fluxos ao descrever o que deseja."],
    ["Como evitar erros em mudancas?", "Consulte 'Mudancas Seguras' antes de alterar qualquer coisa. Verifique tabelas e regras impactadas."],
    ["O que significam as cores?", "Azul escuro = cabecalho. Verde = OK/funcionando. Amarelo = atencao. Vermelho = problema/risco critico."],
    ["Como manter atualizado?", "A cada mudanca no sistema, atualize as abas relevantes. Este arquivo e uma documentacao VIVA."],
]
add_sheet(wb, "00-Guia de Leitura", headers, data)

# =====================================================
# ABA 1: VISAO GERAL DO SISTEMA
# =====================================================
headers = ["Item", "Descricao"]
data = [
    ["Nome do Sistema", "Sistema Financeiro Multi-Tenant para BPO Contabil"],
    ["Objetivo", "Gerenciar contas a pagar/receber, conciliacao bancaria, estoque, notas fiscais e relatorios financeiros para multiplas empresas"],
    ["Stack Tecnologico", "Next.js 16, React 19, Prisma ORM, PostgreSQL (Neon), Vercel, shadcn/ui, Tailwind CSS"],
    ["Autenticacao", "NextAuth v5 com JWT + Credentials Provider (email/senha com bcryptjs)"],
    ["Banco de Dados", "PostgreSQL hospedado no Neon (sa-east-1), 35 tabelas, 23 enums"],
    ["---", "--- MODULOS PRINCIPAIS ---"],
    ["Modulo: Dashboard", "Visao geral com KPIs, graficos de receita vs despesa, fluxo de caixa, aging. Dashboard multi-tenant para visao consolidada."],
    ["Modulo: Cadastros", "Plano de Contas, Fornecedores, Clientes, Centros de Custo, Contas Bancarias, Formas de Pagamento, Produtos, Depositos"],
    ["Modulo: Importacoes", "Extrato bancario (OFX/CSV), Transacoes de cartao, Notas fiscais (XML/CSV), Notas de compra, Pluggy (Open Banking), QIVE (NFe automatica)"],
    ["Modulo: Staging (Pre-lancamento)", "Area intermediaria onde importacoes sao classificadas, validadas e incorporadas como lancamentos oficiais"],
    ["Modulo: Financeiro", "Lancamentos oficiais (contas a pagar/receber), liquidacao (baixa), parcelamento, calendario financeiro, regras recorrentes"],
    ["Modulo: Conciliacao", "Conciliacao bancaria (automatica e manual) e conciliacao de cartoes"],
    ["Modulo: Estoque", "Movimentacoes, Kardex por produto, Posicao de estoque"],
    ["Modulo: Relatorios", "DRE, Aging, Orcado vs Realizado, Fluxo de Caixa"],
    ["Modulo: Controles", "Check diario, Log de auditoria, Excecoes, Checklist de fechamento"],
    ["Modulo: Configuracoes", "Empresas, Usuarios/Permissoes, Regras de classificacao, Periodos bloqueados, Onboarding"],
    ["---", "--- USUARIOS PRINCIPAIS ---"],
    ["ADMIN", "Administrador - acesso total, cria empresas, convida usuarios, aplica templates"],
    ["CONTROLLER", "Controller - incorpora lancamentos, sincroniza integracoes, gerencia dados financeiros"],
    ["ANALYST", "Analista - opera o dia a dia, importa dados, classifica lancamentos"],
    ["VIEWER", "Visualizador - acesso somente leitura a dashboards e relatorios"],
    ["---", "--- FLUXO PRINCIPAL ---"],
    ["1. Entrada", "Dados entram via importacao (OFX, CSV, XML, Pluggy, QIVE) ou cadastro manual"],
    ["2. Staging", "Entradas ficam na area de staging para classificacao (automatica ou manual)"],
    ["3. Validacao", "Entradas classificadas sao validadas (verifica campos, duplicatas, periodo bloqueado)"],
    ["4. Incorporacao", "Entradas validadas viram lancamentos oficiais com numero sequencial"],
    ["5. Liquidacao", "Lancamentos sao baixados (parcial ou total) com atualizacao de saldo bancario"],
    ["6. Conciliacao", "Lancamentos liquidados sao conciliados com extrato bancario"],
    ["---", "--- INTEGRACOES ---"],
    ["Pluggy", "Open Banking - sincroniza extratos bancarios automaticamente via API"],
    ["QIVE", "Sincroniza notas fiscais eletronicas (NFe) automaticamente"],
    ["Cron Jobs", "2 jobs: geracao de lancamentos recorrentes + deteccao de vencidos"],
    ["---", "--- RISCOS OPERACIONAIS ---"],
    ["R1: Sem verificacao de periodo na baixa", "Lancamentos podem ser liquidados em periodos ja fechados"],
    ["R2: Cancelamento sem checagem de tenant", "Falha de seguranca permite cancelar lancamentos de outra empresa"],
    ["R3: Webhook Pluggy sem autenticacao", "Endpoint aceita qualquer requisicao POST sem verificar assinatura"],
    ["R4: Sem paginacao nas listagens", "Telas limitadas a 200-500 registros, sem paginacao real"],
    ["R5: Race condition no numero sequencial", "Incorporacoes simultaneas podem gerar numeros duplicados"],
]
add_sheet(wb, "01-Visao Geral", headers, data)

# =====================================================
# ABA 3: TABELAS DO BANCO / ENTIDADES
# =====================================================
headers = ["Tabela", "Nome Amigavel", "Objetivo", "Modulo", "Tipo", "Chave Primaria", "Chaves Estrangeiras", "Principais Relacionamentos", "Frequencia Atualizacao", "Sensibilidade", "Obs"]
data = [
    ["Tenant", "Empresa", "Entidade raiz multi-tenant, representa uma empresa", "Admin", "Cadastro", "id (cuid)", "Nenhuma", "1:N com quase todas as tabelas", "Rara", "Alta", "Soft-delete via campo active"],
    ["User", "Usuario", "Identidade do usuario autenticado", "Auth", "Cadastro", "id (cuid)", "Nenhuma", "1:N com Membership, Session, Account, varias tabelas via createdById", "Rara", "Alta", "Senha hasheada com bcryptjs"],
    ["Membership", "Vinculo Usuario-Empresa", "Associa usuario a empresa com papel", "Auth", "Apoio", "id (cuid)", "userId->User, tenantId->Tenant", "N:N entre User e Tenant", "Rara", "Media", "Unique: userId+tenantId"],
    ["Account", "Conta OAuth", "Armazena tokens de provedores OAuth (NextAuth)", "Auth", "Apoio", "id (cuid)", "userId->User", "1:N User->Account", "Rara", "Alta", "Tokens de acesso sensiveis"],
    ["Session", "Sessao", "Sessoes de usuario (NextAuth)", "Auth", "Apoio", "id (cuid)", "userId->User", "1:N User->Session", "Frequente", "Alta", "Nao usado ativamente (JWT)"],
    ["VerificationToken", "Token Verificacao", "Tokens para verificacao de email", "Auth", "Apoio", "Composto", "Nenhuma", "Nenhum", "Rara", "Alta", "Sem PK explicita"],
    ["ChartOfAccount", "Plano de Contas", "Estrutura hierarquica de contas contabeis", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant, parentId->self", "1:N com OfficialEntry, StagingEntry, BudgetLine, RecurringRule, ClassificationRule", "Ocasional", "Media", "Auto-referencial (pai/filho). Unique: tenantId+code"],
    ["Supplier", "Fornecedor", "Cadastro de fornecedores", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com PurchaseInvoice, StagingEntry, OfficialEntry, ClassificationRule", "Ocasional", "Media", "Unique: tenantId+cnpjCpf"],
    ["Customer", "Cliente", "Cadastro de clientes", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com StagingEntry, OfficialEntry, ClassificationRule", "Ocasional", "Media", "Unique: tenantId+cnpjCpf"],
    ["CostCenter", "Centro de Custo", "Classificacao gerencial de despesas/receitas", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant, parentId->self", "1:N com OfficialEntry, StagingEntry, BudgetLine, RecurringRule", "Rara", "Baixa", "Auto-referencial. Unique: tenantId+code"],
    ["BankAccount", "Conta Bancaria", "Contas bancarias da empresa", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com BankStatementLine, OfficialEntry, Settlement, RecurringRule, PluggyConnection", "Diaria (saldo)", "Alta", "currentBalance atualizado a cada liquidacao"],
    ["PaymentMethod", "Forma de Pagamento", "Formas de pagamento disponiveis", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com StagingEntry, OfficialEntry, Settlement, RecurringRule", "Rara", "Baixa", "feePercentage nunca usado na liquidacao (BUG)"],
    ["Product", "Produto", "Catalogo de produtos/servicos", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com PurchaseInvoiceItem, StockMovement", "Ocasional", "Baixa", "Unique: tenantId+code"],
    ["Warehouse", "Deposito", "Locais de armazenamento", "Cadastros", "Cadastro", "id (cuid)", "tenantId->Tenant", "1:N com StockMovement", "Rara", "Baixa", ""],
    ["ClassificationRule", "Regra de Classificacao", "Regras para classificacao automatica de lancamentos", "Config", "Parametrizacao", "id (cuid)", "tenantId->Tenant, chartOfAccountId->ChartOfAccount, costCenterId->CostCenter, supplierId->Supplier, customerId->Customer", "Aplicada em StagingEntry", "Ocasional", "Baixa", "Prioridade define ordem de avaliacao"],
    ["PeriodLock", "Periodo Bloqueado", "Controle de fechamento mensal contabil", "Config", "Parametrizacao", "id (cuid)", "tenantId->Tenant, lockedById->User", "Consultado na validacao de staging", "Mensal", "Media", "Unique: tenantId+year+month. NAO verificado na liquidacao (BUG)"],
    ["ImportBatch", "Lote de Importacao", "Cabecalho de cada importacao de dados", "Importacoes", "Transacao", "id (cuid)", "tenantId->Tenant, importedById->User", "1:N com TaxInvoiceLine, BankStatementLine, CardTransaction, PurchaseInvoice, StagingEntry", "Diaria", "Baixa", "Status: PENDING->PROCESSING->COMPLETED/FAILED"],
    ["TaxInvoiceLine", "Linha de NF", "Itens de notas fiscais importadas (NF-e)", "Importacoes", "Transacao", "id (cuid)", "tenantId->Tenant, importBatchId->ImportBatch", "Origem de StagingEntry", "Diaria", "Media", "Unique: tenantId+externalId"],
    ["BankStatementLine", "Linha de Extrato", "Linhas do extrato bancario importado", "Importacoes", "Transacao", "id (cuid)", "tenantId->Tenant, importBatchId->ImportBatch, bankAccountId->BankAccount", "1:1 com Reconciliation", "Diaria", "Media", "Unique: tenantId+bankAccountId+externalId"],
    ["CardTransaction", "Transacao de Cartao", "Transacoes de cartao de credito/debito importadas", "Importacoes", "Transacao", "id (cuid)", "tenantId->Tenant, importBatchId->ImportBatch", "Origem de StagingEntry", "Diaria", "Media", ""],
    ["PurchaseInvoice", "Nota de Compra", "Cabecalho de notas de entrada (compra)", "Importacoes", "Transacao", "id (cuid)", "tenantId->Tenant, importBatchId->ImportBatch, supplierId->Supplier", "1:N com PurchaseInvoiceItem", "Diaria", "Media", ""],
    ["PurchaseInvoiceItem", "Item de Nota de Compra", "Itens de notas de compra", "Importacoes", "Transacao", "id (cuid)", "purchaseInvoiceId->PurchaseInvoice, productId->Product", "Vincula NF a Produto", "Diaria", "Baixa", ""],
    ["StagingEntry", "Lancamento Provisorio", "Area intermediaria antes de virar lancamento oficial", "Staging", "Transacao", "id (cuid)", "tenantId->Tenant, importBatchId->ImportBatch + 8 FKs opcionais", "1:1 opcional com OfficialEntry (via stagingEntryId)", "Diaria", "Media", "Status: PENDING->AUTO_CLASSIFIED->VALIDATED->INCORPORATED. Faltam campos dueDate e competenceDate (BUG)"],
    ["OfficialEntry", "Lancamento Oficial", "Lancamento financeiro definitivo - nucleo do sistema", "Financeiro", "Transacao", "id (cuid)", "tenantId->Tenant + 9 FKs", "1:N com Settlement, Reconciliation. 1:1 com StagingEntry", "Diaria", "Alta", "Status: OPEN->PARTIAL->SETTLED/CANCELLED. Unique: tenantId+sequentialNumber"],
    ["Settlement", "Liquidacao (Baixa)", "Registro de pagamento/recebimento", "Financeiro", "Transacao", "id (cuid)", "tenantId->Tenant, officialEntryId->OfficialEntry, bankAccountId->BankAccount, paymentMethodId->PaymentMethod, settledById->User", "1:1 opcional com Reconciliation", "Diaria", "Alta", "Atualiza saldo da conta bancaria. Sem estorno (BUG)"],
    ["Reconciliation", "Conciliacao", "Match entre extrato e lancamento/liquidacao", "Conciliacao", "Transacao", "id (cuid)", "tenantId->Tenant, bankStatementLineId->BankStatementLine, officialEntryId->OfficialEntry, settlementId->Settlement, reconciledById->User", "Liga extrato a liquidacao", "Diaria", "Media", "Unique: bankStatementLineId e settlementId"],
    ["RecurringRule", "Regra Recorrente", "Configuracao de lancamentos recorrentes", "Financeiro", "Parametrizacao", "id (cuid)", "tenantId->Tenant + 7 FKs", "1:N com OfficialEntry gerados", "Conforme frequencia", "Baixa", "Cron job processa regras ativas diariamente"],
    ["BudgetLine", "Linha de Orcamento", "Valores orcados por conta/mes", "Relatorios", "Parametrizacao", "id (cuid)", "tenantId->Tenant, chartOfAccountId->ChartOfAccount, costCenterId->CostCenter", "Comparado com OfficialEntry no relatorio Orcado vs Realizado", "Mensal", "Baixa", "Unique: tenantId+year+month+chartOfAccountId+costCenterId"],
    ["ClosingChecklist", "Checklist Fechamento", "Itens do checklist de fechamento mensal", "Controles", "Parametrizacao", "id (cuid)", "tenantId->Tenant, completedById->User", "Nenhum direto", "Mensal", "Baixa", "Status: PENDING/IN_PROGRESS/COMPLETED/NOT_APPLICABLE"],
    ["StockMovement", "Movimentacao de Estoque", "Entradas, saidas e ajustes de estoque", "Estoque", "Transacao", "id (cuid)", "tenantId->Tenant, productId->Product, warehouseId->Warehouse, createdById->User", "Liga produto a deposito", "Diaria", "Media", "Tipos: ENTRY, EXIT, ADJUSTMENT, TRANSFER"],
    ["AuditLog", "Log de Auditoria", "Registro de todas as alteracoes no sistema", "Controles", "Log/Auditoria", "id (cuid)", "tenantId->Tenant", "Referencia logica via tableName+recordId", "Toda operacao de escrita", "Alta", "Armazena JSON com valores anteriores e novos"],
    ["PluggyConnection", "Conexao Pluggy", "Vinculo com Open Banking via Pluggy", "Integracoes", "Integracao", "id (cuid)", "tenantId->Tenant, bankAccountId->BankAccount, createdById->User", "Gera BankStatementLine e StagingEntry", "Diaria (sync)", "Alta", "Unique: pluggyItemId"],
    ["QiveConnection", "Conexao QIVE", "Vinculo com plataforma QIVE para NFe", "Integracoes", "Integracao", "id (cuid)", "tenantId->Tenant, createdById->User", "Gera TaxInvoiceLine e StagingEntry", "Diaria (sync)", "Alta", "Credenciais via env vars (nao por tenant)"],
    ["TenantInvite", "Convite de Empresa", "Convites para usuarios entrarem em uma empresa", "Admin", "Apoio", "id (cuid)", "tenantId->Tenant, createdById->User", "Cria Membership ao aceitar", "Ocasional", "Media", "Token unico, expira em 7 dias. Status: PENDING/ACCEPTED/EXPIRED/CANCELLED"],
    ["Notification", "Notificacao", "Alertas e avisos para usuarios", "Admin", "Apoio", "id (cuid)", "userId->User", "Nenhum direto", "Diaria", "Baixa", "Tipos: STAGING_PENDING, OVERDUE_PAYABLE, OVERDUE_RECEIVABLE, IMPORT_COMPLETED, INVITE_RECEIVED, SYSTEM"],
]
add_sheet(wb, "03-Tabelas Entidades", headers, data)

# =====================================================
# ABA 4: RELACIONAMENTOS
# =====================================================
headers = ["Tabela Origem", "Campo Origem", "Tabela Destino", "Campo Destino", "Tipo", "Descricao", "Regra Integridade", "Impacto Exclusao", "Obs"]
rels = [
    ["Tenant", "id", "Membership", "tenantId", "1:N", "Empresa tem muitos vinculos de usuario", "CASCADE", "Exclui todos os vinculos", ""],
    ["Tenant", "id", "ChartOfAccount", "tenantId", "1:N", "Empresa tem seu plano de contas", "CASCADE", "Exclui todo plano de contas", ""],
    ["Tenant", "id", "Supplier", "tenantId", "1:N", "Empresa tem seus fornecedores", "CASCADE", "Exclui todos fornecedores", ""],
    ["Tenant", "id", "Customer", "tenantId", "1:N", "Empresa tem seus clientes", "CASCADE", "Exclui todos clientes", ""],
    ["Tenant", "id", "CostCenter", "tenantId", "1:N", "Empresa tem seus centros de custo", "CASCADE", "Exclui todos centros de custo", ""],
    ["Tenant", "id", "BankAccount", "tenantId", "1:N", "Empresa tem suas contas bancarias", "CASCADE", "Exclui todas contas", ""],
    ["Tenant", "id", "StagingEntry", "tenantId", "1:N", "Empresa tem lancamentos provisorios", "CASCADE", "Exclui todos staging", ""],
    ["Tenant", "id", "OfficialEntry", "tenantId", "1:N", "Empresa tem lancamentos oficiais", "CASCADE", "Exclui todos lancamentos", "Impacto muito alto"],
    ["Tenant", "id", "Settlement", "tenantId", "1:N", "Empresa tem liquidacoes", "CASCADE", "Exclui todas liquidacoes", ""],
    ["Tenant", "id", "AuditLog", "tenantId", "1:N", "Empresa tem logs de auditoria", "CASCADE", "Exclui todo historico de auditoria", ""],
    ["User", "id", "Membership", "userId", "1:N", "Usuario participa de varias empresas", "CASCADE", "Remove todos os vinculos do usuario", ""],
    ["User", "id", "Notification", "userId", "1:N", "Usuario recebe notificacoes", "CASCADE", "Remove todas notificacoes", ""],
    ["ChartOfAccount", "id", "ChartOfAccount", "parentId", "1:N", "Conta pai -> contas filhas (hierarquia)", "SetNull", "Filhos ficam sem pai", "Auto-referencial"],
    ["ChartOfAccount", "id", "OfficialEntry", "chartOfAccountId", "1:N", "Conta contabil classifica lancamentos", "Restrict", "Nao permite excluir conta com lancamentos", "Campo obrigatorio no OfficialEntry"],
    ["CostCenter", "id", "CostCenter", "parentId", "1:N", "Centro de custo pai -> filhos", "SetNull", "Filhos ficam sem pai", "Auto-referencial"],
    ["BankAccount", "id", "BankStatementLine", "bankAccountId", "1:N", "Conta bancaria tem linhas de extrato", "Restrict", "Nao permite excluir conta com extratos", ""],
    ["BankAccount", "id", "OfficialEntry", "bankAccountId", "1:N", "Conta bancaria associada a lancamentos", "Restrict", "Nao permite excluir conta com lancamentos", "Campo obrigatorio"],
    ["BankAccount", "id", "Settlement", "bankAccountId", "1:N", "Liquidacoes vinculadas a conta bancaria", "Restrict", "Nao permite excluir conta com liquidacoes", "Atualiza saldo"],
    ["Supplier", "id", "OfficialEntry", "supplierId", "1:N", "Fornecedor vinculado a lancamentos", "SetNull", "Lancamento perde referencia ao fornecedor", "Campo opcional"],
    ["Customer", "id", "OfficialEntry", "customerId", "1:N", "Cliente vinculado a lancamentos", "SetNull", "Lancamento perde referencia ao cliente", "Campo opcional"],
    ["ImportBatch", "id", "TaxInvoiceLine", "importBatchId", "1:N", "Lote agrupa linhas de NF importadas", "CASCADE", "Exclui todas linhas do lote", ""],
    ["ImportBatch", "id", "BankStatementLine", "importBatchId", "1:N", "Lote agrupa linhas de extrato", "CASCADE", "Exclui todas linhas do lote", ""],
    ["ImportBatch", "id", "StagingEntry", "importBatchId", "1:N", "Lote originou lancamentos provisorios", "SetNull", "Staging perde referencia ao lote", ""],
    ["StagingEntry", "id", "OfficialEntry", "stagingEntryId", "1:1", "Staging vira lancamento oficial na incorporacao", "Unique FK", "Lancamento oficial perde rastreabilidade", "stagingEntryId e UNIQUE no OfficialEntry"],
    ["OfficialEntry", "id", "Settlement", "officialEntryId", "1:N", "Lancamento recebe uma ou mais liquidacoes (parciais)", "Restrict", "Nao permite excluir lancamento com liquidacao", ""],
    ["OfficialEntry", "id", "Reconciliation", "officialEntryId", "1:N", "Lancamento pode ser conciliado", "SetNull", "Conciliacao perde referencia", ""],
    ["BankStatementLine", "id", "Reconciliation", "bankStatementLineId", "1:1", "Linha de extrato conciliada", "Unique FK", "Conciliacao excluida junto", "Cada linha so concilia uma vez"],
    ["Settlement", "id", "Reconciliation", "settlementId", "1:1", "Liquidacao conciliada com extrato", "Unique FK", "Conciliacao excluida junto", "Cada liquidacao so concilia uma vez"],
    ["RecurringRule", "id", "OfficialEntry", "recurringRuleId", "1:N", "Regra gera lancamentos periodicamente", "SetNull", "Lancamentos gerados permanecem", ""],
    ["Product", "id", "StockMovement", "productId", "1:N", "Produto tem movimentacoes de estoque", "Restrict", "Nao permite excluir produto com movimentacoes", ""],
    ["Product", "id", "PurchaseInvoiceItem", "productId", "1:N", "Produto em itens de nota de compra", "Restrict", "Nao permite excluir produto com notas", ""],
    ["Warehouse", "id", "StockMovement", "warehouseId", "1:N", "Deposito registra movimentacoes", "Restrict", "Nao permite excluir deposito com movimentacoes", ""],
    ["PurchaseInvoice", "id", "PurchaseInvoiceItem", "purchaseInvoiceId", "1:N", "Nota de compra tem itens", "CASCADE", "Exclui todos itens da nota", ""],
]
add_sheet(wb, "04-Relacionamentos", headers, rels)

# =====================================================
# ABA 5: REGRAS DE NEGOCIO
# =====================================================
headers = ["ID", "Nome", "Descricao Simples", "Descricao Tecnica", "Modulo", "Processo", "Condicao Entrada", "Validacao", "Acao Executada", "Resultado", "Excecoes", "Risco se Falhar", "Prioridade", "Obs"]
rules = [
    ["RN01", "Validacao de Staging", "Lancamento provisorio precisa ter campos minimos preenchidos", "validateStagingEntry: verifica chartOfAccountId, bankAccountId, date, description, amount>0", "Staging", "Validacao de lancamento", "Status=PENDING ou AUTO_CLASSIFIED", "5 checagens de campo + duplicata + periodo bloqueado", "Status muda para VALIDATED", "Lancamento pronto para incorporacao", "Campos opcionais nao bloqueiam", "Lancamento invalido e incorporado sem classificacao", "Critica", ""],
    ["RN02", "Deteccao de Duplicata", "Sistema verifica se ja existe lancamento igual", "Busca outro StagingEntry com mesma date+amount+description e status!=REJECTED", "Staging", "Validacao", "Lancamento em validacao", "Query por date+amount+description no mesmo tenant", "Retorna warning (nao bloqueia)", "Alerta de possivel duplicata", "Nao filtra por tipo (CREDIT/DEBIT) ou bankAccount", "Duplicatas nao detectadas ou falsos positivos", "Alta", "BUG: nao filtra por tipo nem conta"],
    ["RN03", "Verificacao de Periodo", "Nao permite lancamento em periodo ja fechado", "Consulta PeriodLock por tenantId+year+month", "Staging", "Validacao", "Lancamento em validacao", "Se existe PeriodLock para o mes/ano, bloqueia", "Erro retornado na validacao", "Lancamento nao pode ser validado", "Nenhuma", "Lancamentos em periodos fechados", "Critica", "NAO aplicada na liquidacao (BUG)"],
    ["RN04", "Incorporacao de Staging", "Lancamentos validados viram lancamentos oficiais", "incorporateStagingEntries: cria OfficialEntry, muda status para INCORPORATED", "Staging->Financeiro", "Incorporacao", "Status=VALIDATED", "Somente VALIDATED", "Cria OfficialEntry com numero sequencial, atualiza staging", "Lancamento oficial criado", "Nenhuma", "Lancamento perdido entre staging e oficial", "Critica", "Requer ADMIN ou CONTROLLER"],
    ["RN05", "Numero Sequencial", "Cada lancamento oficial recebe numero unico sequencial", "Max(sequentialNumber)+1 por tenant", "Financeiro", "Incorporacao", "Criacao de OfficialEntry", "Busca ultimo numero e incrementa", "Proximo numero atribuido", "Rastreabilidade sequencial", "Race condition em incorporacoes simultaneas", "Numeros duplicados ou gaps", "Critica", "BUG: nao e atomico"],
    ["RN06", "Categoria Derivada", "Sistema determina se e pagar/receber pelo tipo", "DEBIT->PAYABLE, CREDIT->RECEIVABLE", "Staging->Financeiro", "Incorporacao", "Tipo do lancamento", "Mapeamento fixo tipo->categoria", "Categoria definida automaticamente", "Conta a pagar ou receber", "Transferencias e ajustes nao suportados", "Categoria errada", "Alta", "BUG: deveria usar ChartOfAccount.type"],
    ["RN07", "Liquidacao (Baixa)", "Registrar pagamento/recebimento contra lancamento", "settleEntry: cria Settlement, atualiza OfficialEntry e BankAccount", "Financeiro", "Liquidacao", "OfficialEntry com status OPEN ou PARTIAL", "Valor nao pode exceder saldo restante (tolerancia 0.01)", "Cria Settlement, atualiza status e saldo", "SETTLED (total) ou PARTIAL", "Tolerancia de arredondamento", "Saldo bancario incorreto, status errado", "Critica", ""],
    ["RN08", "Atualizacao de Saldo", "Saldo da conta bancaria e atualizado na liquidacao", "PAYABLE: saldo diminui. RECEIVABLE: saldo aumenta. Juros/multa somam, desconto subtrai", "Financeiro", "Liquidacao", "Liquidacao processada", "Nenhuma alem da propria liquidacao", "currentBalance atualizado via increment/decrement", "Saldo reflete movimentacao real", "Nenhuma", "Saldo bancario nao bate com real", "Critica", ""],
    ["RN09", "Parcelamento", "Dividir lancamento em N parcelas", "generateInstallments: cancela original, cria N novos com installmentGroupId", "Financeiro", "Parcelamento", "OfficialEntry com status OPEN", "Original deve estar OPEN", "Original cancelado, N novos criados", "Parcelas com vencimentos escalonados", "Nenhuma", "Valor total diverge do original", "Alta", "Arredondamento: ultimo recebe diferenca"],
    ["RN10", "Classificacao Automatica", "Sistema classifica lancamentos usando regras", "autoClassify: avalia regras por prioridade, match por CNPJ/descricao/valor", "Staging", "Classificacao", "StagingEntry com status PENDING", "Regras avaliadas em ordem de prioridade", "Campos de classificacao preenchidos, status->AUTO_CLASSIFIED", "Lancamento pre-classificado", "Nenhuma regra corresponde -> continua PENDING", "Classificacao manual necessaria em todos", "Media", "3 tipos de match: CNPJ, DESCRIPTION, VALUE_RANGE"],
    ["RN11", "Conciliacao Automatica", "Match extrato bancario com liquidacoes", "autoReconcile: 3 passes - exato, tolerancia data, tolerancia valor", "Conciliacao", "Conciliacao Bancaria", "BankStatementLine e Settlement nao conciliados", "Pass1: data+valor exatos. Pass2: +/-2 dias. Pass3: valor +/-1%", "Cria Reconciliation com matchType", "Linhas de extrato conciliadas", "Nenhuma correspondencia -> permanece aberta", "Itens nao conciliados se acumulam", "Media", ""],
    ["RN12", "Lancamentos Recorrentes", "Gerar lancamentos automaticamente conforme regra", "Cron: busca RecurringRule ativas com nextGenerationDate<=hoje", "Financeiro", "Cron Job", "RecurringRule ativa, data atingida", "nextGenerationDate <= hoje, active=true, endDate nao expirada", "Cria StagingEntry, atualiza nextGenerationDate", "Novo lancamento no staging", "Se endDate passa, desativa regra", "Lancamento nao gerado, cobranca esquecida", "Alta", "Cria no Staging, NAO no OfficialEntry direto"],
    ["RN13", "Deteccao de Vencidos", "Identificar lancamentos com vencimento ultrapassado", "Cron: busca OfficialEntry com dueDate<hoje e status OPEN/PARTIAL", "Financeiro", "Cron Job", "Execucao diaria do cron", "dueDate < hoje, status IN (OPEN, PARTIAL)", "Cria AuditLog registrando deteccao", "Registro de atraso no log", "Pode gerar duplicatas se rodar varias vezes", "Vencidos nao detectados", "Media", "NAO cria Notification (BUG). NAO tem status OVERDUE"],
    ["RN14", "Soft Delete de Cadastros", "Exclusao logica de registros mestres", "Campos active=false em vez de DELETE fisico", "Cadastros", "Exclusao", "Registro existe e esta ativo", "Nenhuma validacao de dependencias", "Campo active muda para false", "Registro inativo, nao aparece em listagens", "Nenhuma", "Dados orfaos se dependencias ativas", "Media", "Aplicado em: Tenant, ChartOfAccount, Supplier, Customer, etc"],
    ["RN15", "Controle de Acesso por Papel", "Acoes restritas por papel do usuario", "requireRole([roles]) verifica memberRole do JWT", "Auth", "Todas operacoes", "Usuario autenticado", "memberRole deve estar na lista permitida", "Operacao executada ou redirecionado", "Seguranca de acesso", "Nem todas acoes verificam papel", "Acesso indevido a funcoes criticas", "Critica", "BUG: varias acoes financeiras nao verificam papel"],
    ["RN16", "Isolamento Multi-Tenant", "Dados de uma empresa nao podem ser acessados por outra", "Toda query filtra por tenantId do JWT", "Todos", "Todas consultas", "Usuario autenticado com tenantId", "WHERE tenantId = user.tenantId", "Dados filtrados por empresa", "Isolamento de dados", "BUGs: cancelEntry, undoReconciliation, rejectStaging nao filtram tenantId", "Vazamento de dados entre empresas", "Critica", "RLS existe mas nao e usado"],
]
add_sheet(wb, "05-Regras de Negocio", headers, rules)

# =====================================================
# ABA 6: FLUXO DA INFORMACAO
# =====================================================
headers = ["ID", "Nome do Fluxo", "Origem", "Forma Entrada", "Tela/Processo", "Validacoes", "Tabelas Impactadas", "Processamento", "Saida Gerada", "Tela Afetada", "Usuario", "Risco", "Obs"]
flows = [
    ["FL01", "Importacao Extrato Bancario", "Arquivo OFX/CSV do banco", "Upload de arquivo", "/imports/bank-statements", "Formato OFX valido, conta bancaria selecionada", "ImportBatch, BankStatementLine, StagingEntry", "Parse do arquivo, criacao de linhas e staging, auto-classificacao", "Lancamentos provisorios no staging", "Staging, Dashboard", "Analyst+", "Medio - sem dedup de linhas ja importadas", "Suporta OFX, CSV e TXT"],
    ["FL02", "Importacao NF-e XML", "Arquivo XML de nota fiscal", "Upload de arquivo", "/imports/tax-invoices", "XML valido com tags NFe", "ImportBatch, TaxInvoiceLine, StagingEntry", "Parse XML, extracao de dados fiscais, criacao de staging", "Lancamentos provisorios + dados fiscais", "Staging, Notas Fiscais", "Analyst+", "Baixo - tem dedup por accessKey", "Suporta multiplos XMLs e multi-NFe"],
    ["FL03", "Importacao Transacoes Cartao", "Arquivo CSV de operadora", "Upload de arquivo", "/imports/card-transactions", "Colunas mapeadas corretamente", "ImportBatch, CardTransaction, StagingEntry", "Parse CSV, criacao de transacoes e staging", "Lancamentos provisorios", "Staging", "Analyst+", "Medio - sem dedup, tudo vira CREDIT", ""],
    ["FL04", "Importacao Notas de Compra", "Arquivo CSV/XLSX", "Upload de arquivo", "/imports/purchase-invoices", "Colunas corretas, CNPJ fornecedor", "ImportBatch, PurchaseInvoice, PurchaseInvoiceItem, StagingEntry", "Parse, auto-vincula fornecedor por CNPJ, cria staging", "NF entrada + staging", "Staging, Fornecedores", "Analyst+", "Alto - supplierId pode ser undefined (BUG)", ""],
    ["FL05", "Sync Pluggy (Open Banking)", "API Pluggy", "Integracao automatica", "/imports/pluggy", "Conexao ativa, conta vinculada", "ImportBatch, BankStatementLine, StagingEntry, PluggyConnection", "Busca transacoes via API, dedup por externalId, cria staging", "Extrato + staging automaticos", "Staging, Dashboard", "Controller+", "Medio - webhook sem autenticacao", "Dedup funciona bem"],
    ["FL06", "Sync QIVE (NFe)", "API QIVE", "Integracao automatica", "/imports/qive", "Credenciais validas, conexao ativa", "ImportBatch, TaxInvoiceLine, StagingEntry, QiveConnection", "Busca NFes via API, dedup por accessKey, cria staging", "NFs + staging automaticos", "Staging, Notas Fiscais", "Controller+", "Baixo - boa dedup", "Paginacao por cursor"],
    ["FL07", "Classificacao Automatica", "StagingEntry com status PENDING", "Automatico pos-importacao", "Interno (service)", "Regras de ClassificationRule ativas", "StagingEntry, ClassificationRule", "Avalia regras por prioridade: CNPJ, descricao, faixa de valor", "Staging classificado (AUTO_CLASSIFIED)", "Staging", "Sistema", "Baixo", "Primeiro match ganha"],
    ["FL08", "Classificacao Manual", "Usuario edita staging", "Manual via tela", "/staging", "Campos de classificacao selecionados", "StagingEntry", "Usuario seleciona conta, centro custo, fornecedor, etc", "Staging classificado manualmente", "Staging", "Analyst+", "Baixo", ""],
    ["FL09", "Validacao de Staging", "StagingEntry classificado", "Acao do usuario na tela", "/staging", "RN01: campos obrigatorios, duplicata, periodo", "StagingEntry, AuditLog", "Verifica 7 condicoes, atualiza status", "VALIDATED ou erros retornados", "Staging", "Analyst+", "Alto - validacao insuficiente pode passar erros", ""],
    ["FL10", "Incorporacao", "StagingEntry validados", "Acao do usuario na tela", "/staging", "Status=VALIDATED, papel ADMIN/CONTROLLER", "StagingEntry, OfficialEntry, AuditLog", "Cria OfficialEntry com seq number, muda staging para INCORPORATED", "Lancamento oficial criado", "Financeiro (Entries, Payables, Receivables)", "Controller+", "Critico - race condition no seq number", "RN04, RN05, RN06"],
    ["FL11", "Liquidacao (Baixa)", "OfficialEntry OPEN/PARTIAL", "Dialog na tela", "/financial/entries", "Valor <= saldo restante, conta bancaria selecionada", "Settlement, OfficialEntry, BankAccount, AuditLog", "Cria Settlement, atualiza status e saldo bancario", "Lancamento liquidado (SETTLED ou PARTIAL)", "Entries, Payables, Receivables, Cash Flow, Dashboard", "Analyst+", "Critico - nao verifica periodo bloqueado", "RN07, RN08"],
    ["FL12", "Parcelamento", "OfficialEntry OPEN", "Dialog na tela", "/financial/entries", "Status=OPEN, N parcelas, data 1o vencimento", "OfficialEntry, AuditLog", "Cancela original, cria N novos com installmentGroupId", "N parcelas com vencimentos escalonados", "Installments, Entries, Calendar", "Analyst+", "Medio - arredondamento pode gerar centavo", "RN09"],
    ["FL13", "Conciliacao Automatica", "BankStatementLine + Settlement", "Botao na tela", "/reconciliation/bank", "Conta bancaria selecionada", "Reconciliation, AuditLog", "3 passes: exato, tolerancia data, tolerancia valor", "Linhas conciliadas com match type", "Reconciliation, Dashboard", "Analyst+", "Medio - performance O(N*M)", "RN11"],
    ["FL14", "Geracao Recorrente (Cron)", "RecurringRule ativas", "Cron job diario", "/api/cron/recurring-entries", "CRON_SECRET, regra ativa, data atingida", "StagingEntry, RecurringRule", "Cria StagingEntry para cada regra devida, atualiza proximo date", "Novos staging entries", "Staging", "Sistema", "Alto - sem transacao, sem isolamento de erro", "RN12"],
    ["FL15", "Deteccao Vencidos (Cron)", "OfficialEntry vencidos", "Cron job diario", "/api/cron/overdue-check", "CRON_SECRET", "AuditLog", "Busca entries com dueDate<hoje e status OPEN/PARTIAL, cria audit log", "Registro de atraso", "Audit Log", "Sistema", "Baixo - apenas registra, nao notifica", "RN13"],
    ["FL16", "Troca de Empresa", "Usuario logado", "Seletor no header", "Layout/Header", "Membership ativa, tenant ativo", "Membership", "Desmarca isDefault atual, marca novo", "Contexto muda para outra empresa", "Todas telas (revalidate)", "Todos", "Baixo", "Efeito imediato via JWT refresh"],
]
add_sheet(wb, "06-Fluxo da Informacao", headers, flows)

wb.save(WB_PATH)
print(f"Parte 1 salva: {WB_PATH}")
print(f"Abas criadas: {wb.sheetnames}")
