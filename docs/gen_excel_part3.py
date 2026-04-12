#!/usr/bin/env python3
"""Mapa Mestre - Parte 3: Abas 13-18 + Extras + Reordenar"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

WB_PATH = os.path.join(os.path.dirname(__file__), "Mapa_Mestre_Sistema_Financeiro.xlsx")
wb = openpyxl.load_workbook(WB_PATH)

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def style_header(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER

def style_data(ws, sr, er, nc):
    for r in range(sr, er + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP; cell.border = THIN_BORDER; cell.font = Font(name="Calibri", size=10)

def auto_width(ws, nc, mx=40):
    for col in range(1, nc + 1):
        w = 10
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value: w = max(w, min(len(str(cell.value)), mx))
        ws.column_dimensions[get_column_letter(col)].width = w + 2

def add_sheet(wb, title, headers, data):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    style_header(ws, len(headers))
    for rd in data: ws.append(rd)
    if data: style_data(ws, 2, len(data) + 1, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))
    return ws

# =====================================================
# ABA 13: PROBLEMAS, RISCOS E FRAGILIDADES
# =====================================================
h = ["ID", "Problema", "Tipo", "Causa", "Impacto", "Risco", "Urgencia", "Area", "Solucao", "Complexidade", "Prioridade", "Obs"]
d = [
    ["P01", "cancelEntry sem checagem de tenantId", "Seguranca", "WHERE filtra so por id, sem tenantId", "Usuario pode cancelar lancamento de outra empresa", "CRITICO", "Imediata", "Financeiro", "Adicionar tenantId no WHERE da query", "Baixa", "P0", "Falha de isolamento multi-tenant"],
    ["P02", "undoReconciliation sem checagem de tenantId", "Seguranca", "DELETE filtra so por id", "Usuario pode excluir conciliacao de outra empresa", "CRITICO", "Imediata", "Conciliacao", "Adicionar tenantId no WHERE", "Baixa", "P0", "Falha de isolamento multi-tenant"],
    ["P03", "rejectStagingEntry sem checagem de tenantId", "Seguranca", "WHERE filtra so por id", "Usuario pode rejeitar staging de outra empresa", "CRITICO", "Imediata", "Staging", "Adicionar tenantId no WHERE", "Baixa", "P0", ""],
    ["P04", "Webhook Pluggy sem autenticacao", "Seguranca", "Endpoint POST aceita qualquer request", "Atacante pode forjar webhooks e manipular dados", "ALTO", "Curto prazo", "Integracoes", "Implementar verificacao HMAC do Pluggy", "Media", "P1", ""],
    ["P05", "VIEWER pode liquidar e cancelar lancamentos", "Seguranca", "Falta requireRole nas actions de liquidacao/cancelamento", "Usuario sem permissao altera dados financeiros", "ALTO", "Curto prazo", "Financeiro", "Adicionar requireRole([ADMIN,CONTROLLER]) nas actions", "Baixa", "P1", ""],
    ["P06", "Race condition no numero sequencial", "Integridade", "Max(seq)+1 fora da transacao do banco", "Incorporacoes simultaneas geram numeros duplicados", "ALTO", "Medio prazo", "Financeiro", "Usar SELECT FOR UPDATE ou sequencia nativa do PG", "Media", "P1", ""],
    ["P07", "competenceDate forcado = date na incorporacao", "Logica Negocio", "Staging nao tem campo competenceDate", "DRE incorreto - competencia nao reflete regime real", "ALTO", "Medio prazo", "Staging/Financeiro", "Adicionar campo competenceDate no StagingEntry", "Media", "P1", "Afeta relatorios contabeis"],
    ["P08", "dueDate forcado = date na incorporacao", "Logica Negocio", "Staging nao tem campo dueDate", "Vencimento sempre = data do lancamento, aging incorreto", "ALTO", "Medio prazo", "Staging/Financeiro", "Adicionar campo dueDate no StagingEntry", "Media", "P1", "Afeta calendar e aging"],
    ["P09", "Categoria derivada simplista (DEBIT->PAYABLE)", "Logica Negocio", "Ignora ChartOfAccount.type", "Transferencias e ajustes classificados errado", "MEDIO", "Medio prazo", "Financeiro", "Derivar categoria do ChartOfAccount.type", "Baixa", "P2", ""],
    ["P10", "Sem verificacao de periodo na liquidacao", "Logica Negocio", "settleEntry nao consulta PeriodLock", "Lancamentos em periodo fechado podem ser liquidados", "ALTO", "Medio prazo", "Financeiro", "Adicionar check PeriodLock em settleEntry", "Baixa", "P1", ""],
    ["P11", "Entries page chama APIs que nao existem", "Bug Funcional", "client.tsx usa fetch(/api/financial/*) que nao foi criado", "Funcoes de liquidar/cancelar/parcelar falham silenciosamente", "ALTO", "Imediata", "Financeiro", "Trocar fetch por server actions", "Media", "P0", ""],
    ["P12", "Importacoes sem deteccao de duplicata", "Integridade", "Reimportacao do mesmo arquivo cria registros duplicados", "Dados duplicados inflam saldos e relatorios", "MEDIO", "Medio prazo", "Importacoes", "Implementar dedup por hash ou campos chave", "Media", "P2", "Exceto XML que tem dedup"],
    ["P13", "Sem paginacao nas listagens", "Performance", "Queries limitadas a 200-500 sem cursor", "Sistema fica lento com volume de dados", "MEDIO", "Medio prazo", "Todos", "Implementar paginacao cursor-based", "Media", "P2", ""],
    ["P14", "Cron recorrentes sem isolamento de erro", "Resiliencia", "Falha em uma regra para todas", "Regras subsequentes nao processadas", "MEDIO", "Curto prazo", "Financeiro", "Wrap try/catch por regra individual", "Baixa", "P1", ""],
    ["P15", "feePercentage de PaymentMethod nunca aplicado", "Funcionalidade", "settleEntry ignora a taxa da forma de pagamento", "Taxas de cartao/boleto nao calculadas", "MEDIO", "Medio prazo", "Financeiro", "Aplicar taxa automaticamente na liquidacao", "Baixa", "P2", ""],
    ["P16", "Sem estorno de liquidacao", "Funcionalidade", "Nao existe funcao de reverter Settlement", "Erro de liquidacao nao pode ser desfeito", "ALTO", "Medio prazo", "Financeiro", "Criar reverseSettlement com rollback de saldo", "Alta", "P1", ""],
    ["P17", "Login sem rate limiting", "Seguranca", "Sem protecao contra brute force", "Senhas podem ser descobertas por tentativa e erro", "MEDIO", "Curto prazo", "Auth", "Implementar rate limiting ou conta lockout", "Media", "P2", ""],
    ["P18", "RLS definido mas nao utilizado", "Seguranca", "withTenantScope existe mas nunca e chamado", "Isolamento depende de filtros manuais (falivel)", "MEDIO", "Longo prazo", "Todos", "Ativar RLS no PostgreSQL ou manter filtros manuais rigorosos", "Alta", "P3", ""],
    ["P19", "Cancelamento nao reverte saldo bancario", "Logica Negocio", "cancelEntry so muda status, nao reverte settlements", "Saldo bancario fica incorreto apos cancelamento", "ALTO", "Curto prazo", "Financeiro", "Reverter settlements associados ou bloquear se ja liquidado", "Media", "P1", ""],
    ["P20", "Audit log nao cobre todas operacoes", "Compliance", "Varias actions nao criam AuditLog", "Alteracoes criticas sem rastreabilidade", "MEDIO", "Medio prazo", "Controles", "Adicionar createAuditLog em todas actions de escrita", "Baixa", "P2", "cancelEntry, reject, undo reconciliation, cron recurring"],
]
ws = add_sheet(wb, "13-Problemas Riscos", h, d)
# Highlight critical rows
for row in range(2, len(d) + 2):
    risk = ws.cell(row=row, column=6).value
    if risk in ("CRITICO", "ALTO"):
        for col in range(1, len(h) + 1):
            ws.cell(row=row, column=col).fill = ALERT_FILL

# =====================================================
# ABA 14: BACKLOG DE MELHORIAS
# =====================================================
h = ["ID", "Melhoria", "Descricao", "Dor que Resolve", "Modulo", "Impacto Negocio", "Impacto Tecnico", "Risco Mudanca", "Dependencias", "Esforco", "Status", "Prioridade", "Obs"]
d = [
    ["M01", "Corrigir isolamento multi-tenant", "Adicionar tenantId em cancelEntry, undoReconciliation, rejectStaging", "Seguranca critica - vazamento entre empresas", "Financeiro/Staging/Conciliacao", "Critico", "Baixo (3 linhas cada)", "Baixo", "Nenhuma", "1h", "Pendente", "P0", "P01+P02+P03"],
    ["M02", "Corrigir chamadas API inexistentes", "Trocar fetch(/api/financial/*) por server actions no entries/client.tsx", "Liquidacao/cancelamento/parcelamento nao funciona", "Financeiro", "Critico", "Medio", "Baixo", "Nenhuma", "4h", "Pendente", "P0", "P11"],
    ["M03", "Adicionar requireRole na liquidacao/cancelamento", "Restringir settleEntry e cancelEntry a ADMIN/CONTROLLER", "VIEWER pode alterar dados financeiros", "Financeiro", "Alto", "Baixo (2 linhas)", "Baixo", "Nenhuma", "30min", "Pendente", "P1", "P05"],
    ["M04", "Verificar periodo na liquidacao", "Adicionar check PeriodLock em settleEntry", "Liquidacao em periodo fechado", "Financeiro", "Alto", "Baixo", "Baixo", "Nenhuma", "1h", "Pendente", "P1", "P10"],
    ["M05", "Adicionar dueDate e competenceDate no Staging", "Novo campo no schema + formularios + incorporacao", "Vencimento e competencia corretos", "Staging/Financeiro", "Alto", "Medio (migration+service+UI)", "Medio", "Migration do banco", "8h", "Pendente", "P1", "P07+P08"],
    ["M06", "Autenticar webhook Pluggy", "Implementar verificacao HMAC no endpoint", "Seguranca do webhook", "Integracoes", "Alto", "Medio", "Baixo", "Documentacao Pluggy", "4h", "Pendente", "P1", "P04"],
    ["M07", "Corrigir race condition seq number", "Usar SELECT FOR UPDATE ou sequence PG", "Numeros sequenciais duplicados", "Financeiro", "Alto", "Medio", "Medio", "Nenhuma", "4h", "Pendente", "P1", "P06"],
    ["M08", "Implementar estorno de liquidacao", "Nova funcao reverseSettlement + UI", "Erro de liquidacao irreversivel", "Financeiro", "Alto", "Alto (novo service+action+UI)", "Medio", "M01", "16h", "Pendente", "P1", "P16"],
    ["M09", "Derivar categoria do ChartOfAccount.type", "EXPENSE/LIABILITY->PAYABLE, REVENUE/ASSET->RECEIVABLE", "Classificacao errada de transferencias/ajustes", "Financeiro", "Medio", "Baixo", "Medio", "Nenhuma", "2h", "Pendente", "P2", "P09"],
    ["M10", "Dedup em importacoes de extrato/cartao", "Verificar hash ou campos-chave antes de inserir", "Dados duplicados em reimportacao", "Importacoes", "Medio", "Medio", "Baixo", "Nenhuma", "8h", "Pendente", "P2", "P12"],
    ["M11", "Paginacao cursor-based nas listagens", "Implementar cursor pagination em todas as list actions", "Performance com volume crescente", "Todos", "Medio", "Alto (muitas telas)", "Baixo", "Nenhuma", "24h", "Pendente", "P2", "P13"],
    ["M12", "Isolamento de erro no cron recorrentes", "Try/catch individual por regra", "Falha em uma regra bloqueia todas", "Financeiro", "Medio", "Baixo", "Baixo", "Nenhuma", "2h", "Pendente", "P1", "P14"],
    ["M13", "Aplicar taxa da forma de pagamento", "Usar feePercentage do PaymentMethod na liquidacao", "Taxas de cartao/boleto nao calculadas", "Financeiro", "Medio", "Baixo", "Baixo", "Nenhuma", "2h", "Pendente", "P2", "P15"],
    ["M14", "Cancelamento reverter saldo/bloquear se liquidado", "Reverter settlements ou impedir cancel de entry liquidado", "Saldo bancario incorreto apos cancelamento", "Financeiro", "Alto", "Medio", "Medio", "M08", "8h", "Pendente", "P1", "P19"],
    ["M15", "Ampliar cobertura de audit log", "Adicionar createAuditLog em todas actions de escrita", "Alteracoes sem rastreabilidade", "Controles", "Medio", "Baixo (repetitivo)", "Baixo", "Nenhuma", "4h", "Pendente", "P2", "P20"],
    ["M16", "Acoes em Contas a Pagar/Receber", "Adicionar liquidar, cancelar, parcelar nas telas payables/receivables", "Telas read-only, obriga ir em entries", "Financeiro", "Medio", "Medio", "Baixo", "M02", "12h", "Pendente", "P3", ""],
    ["M17", "CRUD de regras recorrentes", "Criar, editar, pausar, excluir regras na tela", "Nao da pra gerenciar regras recorrentes", "Financeiro", "Medio", "Medio", "Baixo", "Nenhuma", "12h", "Pendente", "P3", ""],
    ["M18", "Seletor de periodo nos relatorios", "Adicionar filtro de ano/mes no DRE, Aging, Budget", "Relatorios fixos no ano/mes atual", "Relatorios", "Medio", "Baixo", "Baixo", "Nenhuma", "8h", "Pendente", "P3", ""],
    ["M19", "Export CSV/PDF nos relatorios", "Botao de exportacao nas telas de relatorio", "Nao da pra exportar dados", "Relatorios", "Medio", "Medio", "Baixo", "Nenhuma", "16h", "Pendente", "P3", ""],
    ["M20", "Rate limiting no login", "Implementar protecao contra brute force", "Risco de acesso nao autorizado", "Auth", "Medio", "Medio", "Baixo", "Nenhuma", "8h", "Pendente", "P2", "P17"],
]
add_sheet(wb, "14-Backlog Melhorias", h, d)

# =====================================================
# ABA 15: GLOSSARIO
# =====================================================
h = ["Termo Negocio", "Termo Tecnico", "Significado", "Exemplo", "Onde Aparece", "Obs"]
d = [
    ["Empresa", "Tenant", "Entidade que representa uma empresa no sistema multi-tenant", "ABC Comercio Ltda", "Todas as telas, todas as tabelas", "Cada empresa tem dados isolados"],
    ["Usuario", "User", "Pessoa que acessa o sistema", "joao@email.com", "Login, configuracoes", ""],
    ["Papel/Perfil", "Role", "Nivel de permissao do usuario na empresa", "ADMIN, CONTROLLER, ANALYST, VIEWER", "Configuracoes > Usuarios", "Definido por empresa via Membership"],
    ["Vinculo", "Membership", "Associacao de um usuario a uma empresa com um papel", "Joao e CONTROLLER na ABC", "Interno (nao visivel diretamente)", "N:N entre User e Tenant"],
    ["Plano de Contas", "ChartOfAccount", "Estrutura hierarquica de contas contabeis", "1.1.01 - Caixa", "Cadastros > Plano de Contas", "5 tipos: ASSET, LIABILITY, EQUITY, REVENUE, EXPENSE"],
    ["Fornecedor", "Supplier", "Empresa ou pessoa de quem se compra", "Fornecedor XYZ - 12.345.678/0001-90", "Cadastros > Fornecedores", "Vinculado a contas a PAGAR"],
    ["Cliente", "Customer", "Empresa ou pessoa para quem se vende", "Cliente ABC - 98.765.432/0001-10", "Cadastros > Clientes", "Vinculado a contas a RECEBER"],
    ["Centro de Custo", "CostCenter", "Departamento ou projeto para alocacao de despesas", "CC-01 Administrativo", "Cadastros > Centros de Custo", "Hierarquico (pai/filho)"],
    ["Conta Bancaria", "BankAccount", "Conta no banco da empresa", "Banco do Brasil - Ag 1234 CC 56789", "Cadastros > Contas Bancarias", "Saldo atualizado a cada liquidacao"],
    ["Forma de Pagamento", "PaymentMethod", "Meio usado para pagar/receber", "PIX, Boleto, Cartao, Transferencia", "Cadastros > Formas de Pagamento", "Tem taxa (%) e prazo (dias)"],
    ["Lancamento Provisorio", "StagingEntry", "Registro importado que ainda nao e oficial", "Importacao de extrato -> staging", "Staging", "Precisa classificar, validar e incorporar"],
    ["Lancamento Oficial", "OfficialEntry", "Registro financeiro definitivo no sistema", "Pagamento NF 123 - R$ 1.500", "Financeiro > Lancamentos", "Nucleo do sistema financeiro"],
    ["Conta a Pagar", "PAYABLE (EntryCategory)", "Obrigacao financeira - voce deve pagar", "Aluguel do escritorio", "Contas a Pagar", "OfficialEntry com category=PAYABLE"],
    ["Conta a Receber", "RECEIVABLE (EntryCategory)", "Direito financeiro - voce vai receber", "Venda de produto", "Contas a Receber", "OfficialEntry com category=RECEIVABLE"],
    ["Liquidacao / Baixa", "Settlement", "Registro do pagamento/recebimento efetivo", "Pagamento de R$ 500 via PIX", "Lancamentos (acao Liquidar)", "Atualiza saldo bancario e status do lancamento"],
    ["Conciliacao", "Reconciliation", "Match entre extrato bancario e lancamento interno", "Linha extrato R$ 500 = Settlement R$ 500", "Conciliacao Bancaria", "Garante que sistema e banco batem"],
    ["Nota Fiscal (NF-e)", "TaxInvoiceLine", "Documento fiscal eletronico", "NF 12345 - CFOP 5102", "Notas Fiscais", "Importada via XML, CSV ou QIVE"],
    ["Extrato Bancario", "BankStatementLine", "Linha do extrato bancario importado", "15/03 - PIX Recebido - R$ 200", "Conciliacao", "Importado via OFX, CSV ou Pluggy"],
    ["Regra Recorrente", "RecurringRule", "Configuracao para gerar lancamentos automaticamente", "Aluguel mensal R$ 3.000 todo dia 5", "Regras Recorrentes", "Cron gera StagingEntry automaticamente"],
    ["Periodo Bloqueado", "PeriodLock", "Mes/ano que nao permite mais alteracoes", "Jan/2026 bloqueado em 10/02/2026", "Configuracoes > Geral", "Impede validacao de staging no periodo"],
    ["Lote de Importacao", "ImportBatch", "Grupo de registros importados juntos", "Extrato_BB_Marco2026.ofx - 45 registros", "Importacoes", "Rastreia origem dos dados"],
    ["DRE", "Income Statement", "Demonstracao do Resultado do Exercicio", "Receitas - Despesas = Lucro/Prejuizo", "Relatorios > DRE", "Somente lancamentos SETTLED"],
    ["Aging", "Aging Analysis", "Analise de antiguidade de titulos vencidos", "R$ 5.000 vencidos ha 31-60 dias", "Relatorios > Aging", "Faixas: A Vencer, 1-30, 31-60, 61-90, >90 dias"],
    ["Fluxo de Caixa", "Cash Flow", "Projecao de entradas e saidas de dinheiro", "Proximo mes: +R$ 10k entradas, -R$ 8k saidas", "Fluxo de Caixa", "Realizado (passado) + Projetado (futuro)"],
    ["Regra de Classificacao", "ClassificationRule", "Regra automatica para classificar lancamentos", "CNPJ 12345... -> Conta 3.1.01 Aluguel", "Configuracoes > Geral", "Avaliada por prioridade na auto-classificacao"],
    ["Soft Delete", "active = false", "Exclusao logica - registro nao e apagado, apenas desativado", "Fornecedor inativo", "Todos cadastros", "Mantem historico, nao quebra referencias"],
    ["Chave Primaria (PK)", "id @id @default(cuid())", "Identificador unico de cada registro", "cm1abc2de3f4g5h", "Todas tabelas", "Gerado automaticamente, nunca muda"],
    ["Chave Estrangeira (FK)", "campo + @relation", "Campo que referencia registro de outra tabela", "tenantId aponta para Tenant.id", "Maioria das tabelas", "Garante integridade referencial"],
    ["ENUM", "enum NomeEnum { ... }", "Lista fixa de opcoes validas", "EntryStatus: OPEN, PARTIAL, SETTLED, CANCELLED", "Campos de status, tipo, categoria", "Nao aceita valores fora da lista"],
    ["Migration", "prisma migrate dev", "Alteracao estrutural no banco de dados", "Adicionar coluna dueDate na tabela StagingEntry", "Desenvolvimento", "Requer cuidado - pode perder dados"],
    ["Server Action", "use server function", "Funcao que roda no servidor, chamada pelo frontend", "settleOfficialEntry(data)", "Todas telas com acoes", "Substitui APIs tradicionais no Next.js"],
    ["Cron Job", "Scheduled task", "Tarefa automatica executada em horarios programados", "Gerar lancamentos recorrentes todo dia", "API /api/cron/*", "Autenticado por CRON_SECRET"],
]
add_sheet(wb, "15-Glossario", h, d)

# =====================================================
# ABA 16: PADROES DE NOMENCLATURA
# =====================================================
h = ["Elemento", "Padrao", "Exemplo Correto", "Exemplo Incorreto", "Motivo", "Impacto"]
d = [
    ["Tabela (Model)", "PascalCase, singular, ingles", "OfficialEntry, BankAccount", "official_entries, conta_bancaria", "Convencao Prisma + Next.js", "Consistencia no codigo e queries"],
    ["Campo", "camelCase, ingles descritivo", "chartOfAccountId, dueDate, currentBalance", "plano_contas_id, dt_venc, saldo", "Convencao JavaScript/TypeScript", "Legibilidade no codigo"],
    ["Enum", "PascalCase para nome, UPPER_SNAKE para valores", "EntryStatus { OPEN, PARTIAL, SETTLED }", "entry_status { open, partial }", "Convencao TypeScript", "Distincao clara de constantes"],
    ["Chave Estrangeira", "nomeEntidadeId em camelCase", "tenantId, bankAccountId, supplierId", "tenant_id, id_banco, fk_fornecedor", "Padrao Prisma + JS", "Auto-documentacao do relacionamento"],
    ["Rota de Pagina", "kebab-case, hierarquico", "/financial/entries, /master-data/suppliers", "/financeiro/lancamentos, /MasterData/Suppliers", "Convencao URL padrao", "URLs limpas e previsíveis"],
    ["Server Action", "camelCase, verbo + substantivo", "listOfficialEntries, settleEntry, createTenant", "getEntries, doSettle, newTenant", "Clareza de intencao", "API previsivel e documentavel"],
    ["Arquivo de Pagina", "page.tsx (server) + client.tsx (interativo)", "src/app/(app)/staging/page.tsx", "staging.page.tsx, StagingPage.tsx", "Convencao Next.js App Router", "Roteamento automatico"],
    ["Service", "camelCase, acao especifica", "incorporateStagingEntries, settleEntry", "handleStaging, processEntry", "Especificidade", "Codigo auto-documentado"],
    ["Validacao (Schema)", "camelCase + 'Schema' suffix", "officialEntrySchema, supplierSchema", "OfficialEntryValidation, valida_fornecedor", "Padrao Zod + JS", "Identificacao rapida"],
    ["Tipo/Interface", "PascalCase", "SessionUser, ExceptionInfo", "sessionUser, exception_info", "Convencao TypeScript", "Distincao de tipos vs variaveis"],
    ["Constante/Config", "UPPER_SNAKE_CASE", "CRON_SECRET, ROLE_HIERARCHY", "cronSecret, roleHierarchy", "Convencao para constantes", "Identificacao imediata de constantes"],
    ["Pasta de Modulo", "kebab-case", "master-data, bank-statements, cash-flow", "masterData, BankStatements, fluxo_caixa", "Convencao Next.js", "Organizacao uniforme"],
    ["Log/Status", "UPPER_SNAKE_CASE nos enums", "PENDING, AUTO_CLASSIFIED, INCORPORATED", "pending, autoClassified, incorporated", "Padrao de enums", "Consistencia com banco"],
]
add_sheet(wb, "16-Padroes Nomenclatura", h, d)

# =====================================================
# ABA 17: MUDANCAS SEGURAS
# =====================================================
h = ["Item Alterado", "Tipo Alteracao", "Modulos Impactados", "Tabelas Impactadas", "Regras Impactadas", "Telas Impactadas", "Risco", "Teste", "Backup", "Homologacao", "Validacao Usuario", "Obs"]
d = [
    ["Adicionar campo em tabela", "Schema (migration)", "Depende da tabela", "Tabela alvo + relacionadas", "Validacoes Zod, services que usam a tabela", "Formularios e listagens da entidade", "Medio", "Sim - testar CRUD completo", "Sim - dump do banco antes", "Sim", "Sim - confirmar que campo aparece", "Comando: npx prisma migrate dev"],
    ["Remover campo de tabela", "Schema (migration)", "Todos que referenciam o campo", "Tabela alvo", "Todas que usam o campo", "Todas que exibem o campo", "ALTO", "Sim - testar todas funcoes relacionadas", "OBRIGATORIO", "OBRIGATORIO", "OBRIGATORIO", "PERIGO: pode perder dados. Fazer em etapas."],
    ["Alterar regra de negocio", "Codigo (service/action)", "Modulo da regra + dependentes", "Nenhuma (geralmente)", "A propria regra + derivadas", "Depende da regra", "Alto", "Sim - testar cenarios positivos e negativos", "Sim", "Sim", "Sim - validar resultado esperado", "Documentar regra antiga e nova"],
    ["Adicionar nova tela", "Codigo (page + client)", "Modulo da tela", "Somente leitura (geralmente)", "Nenhuma nova", "Nova tela + menu", "Baixo", "Sim - testar renderizacao e dados", "Nao necessario", "Recomendado", "Sim - confirmar usabilidade", ""],
    ["Alterar enum/status", "Schema + codigo", "Todos que usam o enum", "Tabela que contem o campo enum", "Todas que verificam o status", "Todas que exibem/filtram pelo status", "ALTO", "Sim - testar todos os fluxos de status", "OBRIGATORIO", "OBRIGATORIO", "OBRIGATORIO", "Cuidado com dados existentes no banco"],
    ["Alterar logica de liquidacao", "Codigo (settlement service)", "Financeiro, Conciliacao, Dashboard", "Settlement, OfficialEntry, BankAccount", "RN07, RN08", "Entries, Payables, Receivables, Cash Flow", "CRITICO", "OBRIGATORIO - testar parcial, total, multi", "OBRIGATORIO", "OBRIGATORIO", "OBRIGATORIO", "Qualquer erro afeta saldo bancario"],
    ["Adicionar integracao", "Codigo + config", "Integracoes + Staging", "Nova tabela de conexao + staging", "Novas regras de sync", "Nova tela de config + staging", "Medio", "Sim - testar com sandbox da API", "Nao necessario", "Sim", "Sim", "Usar ambiente de teste da API"],
    ["Alterar permissoes/papeis", "Codigo (actions)", "Todos os modulos afetados", "Membership", "RN15", "Todas telas com acoes restritas", "ALTO", "Sim - testar cada papel", "Sim", "OBRIGATORIO", "OBRIGATORIO", "Risco de bloquear usuarios ou liberar demais"],
    ["Corrigir bug de seguranca", "Codigo (actions)", "Modulo do bug", "Nenhuma (geralmente)", "Depende", "Depende", "Baixo (fix) / ALTO (se errar)", "OBRIGATORIO", "Recomendado", "Recomendado", "Nao obrigatorio", "Prioridade maxima, deploy rapido"],
    ["Alterar plano de contas", "Dados (usuario)", "Financeiro, Relatorios, Staging", "ChartOfAccount, ClassificationRule, BudgetLine", "Classificacao, DRE", "Plano de Contas, DRE, staging", "ALTO", "Nao tecnico - mas validar impacto", "Sim - export antes", "Sim", "OBRIGATORIO", "Pode afetar DRE e classificacao automatica"],
]
add_sheet(wb, "17-Mudancas Seguras", h, d)

# =====================================================
# ABA 18: ENUMS DO SISTEMA
# =====================================================
h = ["Enum", "Valor", "Nome Amigavel", "Descricao", "Onde Usado", "Obs"]
d = [
    ["Role", "ADMIN", "Administrador", "Acesso total ao sistema", "User.role, Membership.role", "Nivel 4"],
    ["Role", "CONTROLLER", "Controller", "Gerencia financeira e incorporacoes", "User.role, Membership.role", "Nivel 3"],
    ["Role", "ANALYST", "Analista", "Operacao diaria e importacoes", "User.role, Membership.role", "Nivel 2"],
    ["Role", "VIEWER", "Visualizador", "Somente leitura", "User.role, Membership.role", "Nivel 1"],
    ["AccountType", "ASSET", "Ativo", "Bens e direitos da empresa", "ChartOfAccount.type", ""],
    ["AccountType", "LIABILITY", "Passivo", "Obrigacoes da empresa", "ChartOfAccount.type", ""],
    ["AccountType", "EQUITY", "Patrimonio Liquido", "Diferenca entre ativo e passivo", "ChartOfAccount.type", ""],
    ["AccountType", "REVENUE", "Receita", "Entradas de recursos", "ChartOfAccount.type", ""],
    ["AccountType", "EXPENSE", "Despesa", "Saidas de recursos", "ChartOfAccount.type", ""],
    ["EntryStatus", "OPEN", "Em Aberto", "Lancamento aguardando pagamento/recebimento", "OfficialEntry.status", "Estado inicial"],
    ["EntryStatus", "PARTIAL", "Parcial", "Pagamento/recebimento parcial realizado", "OfficialEntry.status", ""],
    ["EntryStatus", "SETTLED", "Liquidado", "Totalmente pago/recebido", "OfficialEntry.status", "Estado final positivo"],
    ["EntryStatus", "CANCELLED", "Cancelado", "Lancamento cancelado", "OfficialEntry.status", "Estado final negativo"],
    ["EntryCategory", "PAYABLE", "A Pagar", "Obrigacao - a empresa deve pagar", "OfficialEntry.category", "Contas a pagar"],
    ["EntryCategory", "RECEIVABLE", "A Receber", "Direito - a empresa vai receber", "OfficialEntry.category", "Contas a receber"],
    ["EntryCategory", "TRANSFER", "Transferencia", "Movimentacao entre contas", "OfficialEntry.category", ""],
    ["EntryCategory", "ADJUSTMENT", "Ajuste", "Correcao ou ajuste contabil", "OfficialEntry.category", ""],
    ["TransactionType", "CREDIT", "Credito", "Entrada de recursos na conta", "StagingEntry.type, OfficialEntry.type", "Dinheiro entra"],
    ["TransactionType", "DEBIT", "Debito", "Saida de recursos da conta", "StagingEntry.type, OfficialEntry.type", "Dinheiro sai"],
    ["StagingStatus", "PENDING", "Pendente", "Aguardando classificacao", "StagingEntry.status", "Estado inicial"],
    ["StagingStatus", "AUTO_CLASSIFIED", "Auto-Classificado", "Classificado automaticamente por regra", "StagingEntry.status", ""],
    ["StagingStatus", "VALIDATED", "Validado", "Aprovado para incorporacao", "StagingEntry.status", ""],
    ["StagingStatus", "INCORPORATED", "Incorporado", "Ja virou lancamento oficial", "StagingEntry.status", "Estado final positivo"],
    ["StagingStatus", "REJECTED", "Rejeitado", "Descartado pelo usuario", "StagingEntry.status", "Estado final negativo"],
    ["ImportStatus", "PENDING", "Pendente", "Importacao na fila", "ImportBatch.status", ""],
    ["ImportStatus", "PROCESSING", "Processando", "Importacao em andamento", "ImportBatch.status", ""],
    ["ImportStatus", "COMPLETED", "Concluido", "Importacao finalizada com sucesso", "ImportBatch.status", ""],
    ["ImportStatus", "FAILED", "Falhou", "Erro durante a importacao", "ImportBatch.status", ""],
    ["StagingSource", "MANUAL", "Manual", "Criado manualmente pelo usuario", "StagingEntry.source", ""],
    ["StagingSource", "IMPORT_BANK_STATEMENT", "Extrato Bancario", "Importado de extrato", "StagingEntry.source", ""],
    ["StagingSource", "IMPORT_PLUGGY", "Pluggy", "Sincronizado via Pluggy", "StagingEntry.source", ""],
    ["StagingSource", "IMPORT_QIVE", "QIVE", "Sincronizado via QIVE", "StagingEntry.source", ""],
    ["Frequency", "DAILY", "Diario", "Todos os dias", "RecurringRule.frequency", ""],
    ["Frequency", "WEEKLY", "Semanal", "Uma vez por semana", "RecurringRule.frequency", ""],
    ["Frequency", "MONTHLY", "Mensal", "Uma vez por mes", "RecurringRule.frequency", "Mais comum"],
    ["Frequency", "QUARTERLY", "Trimestral", "A cada 3 meses", "RecurringRule.frequency", ""],
    ["Frequency", "ANNUAL", "Anual", "Uma vez por ano", "RecurringRule.frequency", ""],
    ["MatchType", "AUTO_EXACT", "Automatico Exato", "Match perfeito: data + valor identicos", "Reconciliation.matchType", "Score 100"],
    ["MatchType", "AUTO_DATE_TOLERANCE", "Automatico Tolerancia Data", "Valor exato, data +/- 2 dias", "Reconciliation.matchType", "Score 80"],
    ["MatchType", "AUTO_VALUE_TOLERANCE", "Automatico Tolerancia Valor", "Data exata, valor +/- 1%", "Reconciliation.matchType", "Score 60"],
    ["MatchType", "MANUAL", "Manual", "Match feito manualmente pelo usuario", "Reconciliation.matchType", ""],
    ["PaymentMethodType", "PIX", "PIX", "Pagamento instantaneo", "PaymentMethod.type", ""],
    ["PaymentMethodType", "BOLETO", "Boleto", "Boleto bancario", "PaymentMethod.type", ""],
    ["PaymentMethodType", "BANK_TRANSFER", "Transferencia", "TED/DOC bancaria", "PaymentMethod.type", ""],
    ["PaymentMethodType", "CREDIT_CARD", "Cartao Credito", "Pagamento via cartao de credito", "PaymentMethod.type", ""],
    ["PaymentMethodType", "CASH", "Dinheiro", "Pagamento em especie", "PaymentMethod.type", ""],
    ["AuditAction", "CREATE", "Criacao", "Novo registro criado", "AuditLog.action", ""],
    ["AuditAction", "UPDATE", "Alteracao", "Registro existente modificado", "AuditLog.action", ""],
    ["AuditAction", "DELETE", "Exclusao", "Registro removido ou desativado", "AuditLog.action", ""],
    ["Severity", "LOW", "Baixa", "Problema menor, sem urgencia", "ClosingChecklist.severity", ""],
    ["Severity", "MEDIUM", "Media", "Problema moderado", "ClosingChecklist.severity", ""],
    ["Severity", "HIGH", "Alta", "Problema importante", "ClosingChecklist.severity", ""],
    ["Severity", "CRITICAL", "Critica", "Problema urgente que exige acao imediata", "ClosingChecklist.severity", ""],
]
add_sheet(wb, "18-Enums do Sistema", h, d)

# Reorder sheets
desired = [
    "00-Guia de Leitura", "01-Visao Geral", "02-Dicionario de Dados",
    "03-Tabelas Entidades", "04-Relacionamentos", "05-Regras de Negocio",
    "06-Fluxo da Informacao", "07-Modulos Funcionalidades", "08-Telas Interfaces",
    "09-Entradas Process Saidas", "10-Integracoes", "11-Perfis Permissoes",
    "12-Logs Auditoria", "13-Problemas Riscos", "14-Backlog Melhorias",
    "15-Glossario", "16-Padroes Nomenclatura", "17-Mudancas Seguras",
    "18-Enums do Sistema"
]
order = []
for name in desired:
    if name in wb.sheetnames:
        order.append(wb.sheetnames.index(name))
wb.move_sheet("02-Dicionario de Dados", offset=-10)

wb.save(WB_PATH)
print(f"Parte 3 salva: {WB_PATH}")
print(f"Total abas: {len(wb.sheetnames)}")
print(f"Abas: {wb.sheetnames}")
