#!/usr/bin/env python3
"""Parte 4 – Revisão Arquitetural + Modelo de Dados Alvo + Motor de Regras
   + Atualização de Problemas, Backlog e Dicionário."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy

FILE = "/Users/jorgesalsa/Downloads/teste/financeiro/docs/Mapa_Mestre_Sistema_Financeiro.xlsx"
wb = openpyxl.load_workbook(FILE)

# ── estilos ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DATA_FONT   = Font(name="Calibri", size=10)
ALERT_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OK_FILL     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_data(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 55))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def add_sheet(name, headers, data, highlight_col=None, highlight_vals=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_data(ws, r, len(headers))
        if highlight_col and highlight_vals:
            cell_val = ws.cell(row=r, column=highlight_col).value
            if cell_val in highlight_vals.get("alert", []):
                for c2 in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c2).fill = ALERT_FILL
            elif cell_val in highlight_vals.get("warn", []):
                for c2 in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c2).fill = WARN_FILL
            elif cell_val in highlight_vals.get("ok", []):
                for c2 in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c2).fill = OK_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws

# ═══════════════════════════════════════════════════════════════════════
# ABA 19 – REVISÃO ARQUITETURAL
# ═══════════════════════════════════════════════════════════════════════
rev_headers = [
    "ID", "Ponto Crítico", "Área", "Diagnóstico Detalhado",
    "Risco se Não Corrigir", "Severidade", "Evidência",
    "Recomendação", "Entidades Afetadas", "Prioridade de Correção",
    "Esforço Estimado", "Dependências"
]
rev_data = [
    ["RA01", "Ambiguidade de datas e períodos",
     "Dados / Conciliação",
     "O sistema possui múltiplas datas (emissão, vencimento, pagamento, compensação, competência, importação) mas não declara explicitamente qual campo governa cada processo. A conciliação revisada mostrou 868 linhas classificadas como 'fora do período' por usar data de emissão (com horário) em vez de data de pagamento normalizada, impactando R$ 267.724,53.",
     "Conciliação incorreta, DRE distorcido, aging errado, overdue falso, aprovações sobre dados incorretos. Escala exponencial conforme volume cresce.",
     "CRÍTICO",
     "Planilha de conciliação CFO: 868 linhas fora do período, parecer 'não aprovar sem ajuste'",
     "Criar regra mestra de datas: cada processo declara qual campo usa. Implementar competence_date, bank_posted_date, settlement_date como campos distintos e obrigatórios.",
     "OfficialEntry, Settlement, BankStatementLine, ReconciliationMatch",
     "P0 – Imediato",
     "Alto (refatoração de queries + migração de dados)",
     "Nenhuma – é fundação"],

    ["RA02", "Staging sem estados formais de quarentena",
     "Pipeline de Dados",
     "O StagingEntry funciona como área prévia ao oficial, mas não tem ciclo de vida rígido. Faltam estados intermediários formais: importado_bruto → parseado → normalizado → classificado → com_conflito → validado → rejeitado → incorporado. Sem trilha de auditoria obrigatória em cada transição.",
     "Dados passam do staging para oficial sem validação completa. Impossível rastrear em que ponto um dado foi corrompido ou classificado errado.",
     "CRÍTICO",
     "Código atual: StagingEntry tem campo status mas sem máquina de estados formal",
     "Implementar máquina de estados (state machine) com transições explícitas, cada transição gera AuditLog. Bloquear incorporação de itens que não passaram por todos os estados obrigatórios.",
     "StagingEntry, AuditLog",
     "P0 – Imediato",
     "Médio (enum de estados + validação de transições)",
     "RA01 (datas corretas antes de validar)"],

    ["RA03", "Ausência de motor de regras",
     "Classificação / Automação",
     "O sistema depende de campos estáticos para classificação. Não existe motor de regras que execute lógica como: se origem=extrato E descrição contém X → sugerir categoria Y; se valor divergir além de tolerância → bloquear auto-match; se saída classificada como receita → bloquear incorporação.",
     "Sistema permanece dependente de operador experiente. Classificações inconsistentes, auto-match frágil, zero escalabilidade operacional.",
     "ALTO",
     "ClassificationRule existe mas é simples (keyword→categoria). Não há regras compostas, tolerâncias, bloqueios automáticos.",
     "Implementar motor de regras com: condições compostas (AND/OR), ações (classificar, bloquear, alertar, encaminhar), prioridade/peso, score de confiança, log de regra aplicada.",
     "ClassificationRule, ValidationRule, StagingEntry, OfficialEntry",
     "P1 – Curto prazo",
     "Alto (novo módulo)",
     "RA02 (estados de staging definidos)"],

    ["RA04", "Conciliação sem score de confiança",
     "Conciliação",
     "O modelo atual faz matching mas não registra o grau de certeza, a base da decisão, nem a regra que gerou o match. Campos como match_confidence, match_rule_id, match_basis, requires_human_review, review_reason, approved_by, approved_at não existem.",
     "Impossível auditar por que um match foi feito. Matches errados passam sem revisão. Sem dados para melhorar o algoritmo ao longo do tempo.",
     "ALTO",
     "ReconciliationMatch não tem campos de confiança/auditoria. Planilha de conciliação mostra matches múltiplos legítimos sem critério formal.",
     "Adicionar à ReconciliationMatch: match_confidence (0-100), match_rule_id, match_basis (valor/data/doc/descrição/banco), requires_human_review, review_reason, approved_by, approved_at. Criar fila de revisão para matches abaixo do threshold.",
     "ReconciliationMatch, Settlement, BankStatementLine",
     "P1 – Curto prazo",
     "Médio (novos campos + lógica de fila)",
     "RA01 (datas corretas para matching)"],

    ["RA05", "Taxonomia/categorização misturada",
     "Classificação / Dados Mestres",
     "Categorias atuais misturam natureza contábil (RECEITA), gerencial (SALARIO, MATERIA PRIMA), operacional (TRANSFERENCIA ENTRE CONTAS), provisória (DUVIDA) e até nível de detalhe (BEBIDAS). 'DUVIDA' é status de classificação, não categoria. Sem separação formal entre camadas.",
     "Relatórios gerenciais incoerentes, filtros que não fazem sentido, impossível comparar períodos com categorias que mudam de significado.",
     "ALTO",
     "Base de dados: categorias RECEITA, DUVIDA, SALARIO, MATERIA PRIMA, BONUS, TRANSFERENCIA ENTRE CONTAS, BEBIDAS coexistem no mesmo campo",
     "Separar em 4 camadas: (1) tipo_movimento: entrada/saída/transferência/ajuste; (2) natureza_financeira: operacional/não-operacional/financeira/patrimonial; (3) categoria_gerencial: receita/matéria-prima/folha/aluguel/frete; (4) status_classificação: classificado/pendente/dúvida/conflito.",
     "OfficialEntry, StagingEntry, ClassificationRule, Category",
     "P1 – Curto prazo",
     "Alto (reestruturação de dados mestres + migração)",
     "RA03 (motor de regras para nova taxonomia)"],

    ["RA06", "Transferência entre contas sem modelagem própria",
     "Tesouraria / Fluxo de Caixa",
     "Transferências entre contas da mesma empresa são tratadas como entrada+saída normais. Isso infla receitas, despesas, caixa e prejudica conciliação. Não há entidade ou flag que identifique o par entrada/saída como uma transferência.",
     "Fluxo de caixa distorcido, DRE poluído, conciliação tenta casar transferências como se fossem operações normais. Relatórios perdem credibilidade.",
     "ALTO",
     "Categoria 'TRANSFERENCIA ENTRE CONTAS' existe mas é tratada como qualquer outra categoria, sem lógica de pareamento",
     "Criar entidade InternalTransfer com: conta_origem, conta_destino, valor, data, referência. Auto-match por par. Excluir de DRE e relatórios operacionais. Flag transfer_pair_id nos lançamentos.",
     "OfficialEntry, BankAccount, Settlement, ReconciliationMatch",
     "P1 – Curto prazo",
     "Médio (nova entidade + lógica de pareamento)",
     "RA05 (tipo_movimento separado)"],

    ["RA07", "Baixa pode alterar histórico do título",
     "Financeiro / Auditoria",
     "O design precisa garantir que Settlement (baixa) nunca altere o OfficialEntry (título original). O compromisso original deve permanecer imutável. Baixas parciais, juros, multas e descontos devem ser registros separados na Settlement.",
     "Perda de rastreabilidade: impossível saber o valor original de um título, quanto foi pago, quando e por quê. Auditoria comprometida.",
     "MÉDIO",
     "Fluxo de baixa no código atualiza campos do OfficialEntry em alguns cenários",
     "Tornar OfficialEntry imutável após incorporação (exceto status). Toda alteração financeira via Settlement. Campos obrigatórios em Settlement: gross_amount, discount_amount, fine_amount, interest_amount, fee_amount, net_amount, paid_amount, balance_remaining.",
     "OfficialEntry, Settlement, AuditLog",
     "P2 – Médio prazo",
     "Médio (refatoração de fluxo de baixa)",
     "RA02 (estados formais de staging/oficial)"],

    ["RA08", "Parcelamento/recorrência sem versionamento",
     "Financeiro / Contas a Pagar-Receber",
     "Parcelas e recorrências são geradas mas sem snapshot da regra original. Se a regra mudar, parcelas futuras mudam retroativamente. Faltam: origem_mãe, nº parcela, total parcelas, regra de geração, data prevista, flag de edição manual, quem editou, motivo.",
     "Parcelas inconsistentes, impossível auditar por que uma parcela tem aquele valor. Risco de fraude operacional (edição de parcela sem rastro).",
     "MÉDIO",
     "RecurringRule e Installment existem mas sem versionamento/snapshot",
     "Cada geração de parcelas deve criar snapshot imutável da regra. Parcela deve ter: parent_entry_id, installment_number, total_installments, generation_rule_snapshot, expected_date, manually_edited, edited_by, edit_reason.",
     "RecurringRule, Installment, OfficialEntry, AuditLog",
     "P2 – Médio prazo",
     "Médio (campos novos + snapshot logic)",
     "Nenhuma"],

    ["RA09", "Importação não guarda bruto/metadados",
     "Pipeline de Dados / Auditoria",
     "Importações não guardam: hash do arquivo, quantidade de registros, total financeiro importado, erros de parsing, versão do parser. Sem isso, impossível provar de onde veio o dado e se a importação foi completa.",
     "Sem rastreabilidade de origem. Se contestado, não há como provar que o dado veio de determinado arquivo/API. Erros de parsing passam silenciosamente.",
     "MÉDIO",
     "ImportBatch existe mas com metadados limitados",
     "Enriquecer ImportBatch com: file_hash (SHA-256), file_size, record_count, total_amount, parse_errors[], parser_version, source_type (file/api/manual), source_name, raw_file_stored (S3/blob).",
     "ImportBatch, StagingEntry",
     "P2 – Médio prazo",
     "Baixo (campos novos na entidade existente)",
     "Nenhuma"],
]

add_sheet("19-Revisao Arquitetural", rev_headers, rev_data,
          highlight_col=6,
          highlight_vals={"alert": ["CRÍTICO"], "warn": ["ALTO"], "ok": ["MÉDIO"]})

print("✅ Aba 19 criada")

# ═══════════════════════════════════════════════════════════════════════
# ABA 20 – MODELO DE DADOS ALVO (TARGET ARCHITECTURE)
# ═══════════════════════════════════════════════════════════════════════
mdl_headers = [
    "Entidade", "Categoria", "Campo", "Tipo", "Obrigatório",
    "Existe Hoje?", "Descrição Negócio", "Exemplo", "Validação",
    "Observação"
]
mdl_data = [
    # ─── Campos universais (toda entidade financeira) ───
    ["[TODAS]", "Identidade", "id", "UUID/CUID", "SIM", "SIM", "Identificador único interno imutável", "cm3abc123...", "Auto-gerado", "Nunca reutilizar"],
    ["[TODAS]", "Multi-tenant", "tenant_id", "String FK", "SIM", "SIM", "ID da empresa dona do registro", "tenant_abc", "FK → Tenant", "Filtro obrigatório em TODA query"],
    ["[TODAS]", "Rastreabilidade", "source_system", "String", "SIM", "PARCIAL", "Sistema de origem do dado", "pluggy / qive / manual / csv", "Enum", "Saber de onde veio"],
    ["[TODAS]", "Rastreabilidade", "source_record_id", "String", "NÃO", "PARCIAL", "ID do registro no sistema de origem", "txn_12345", "—", "Para conciliar com origem"],
    ["[TODAS]", "Auditoria", "created_at", "DateTime", "SIM", "SIM", "Data/hora de criação", "2026-03-15T10:30:00Z", "Auto", "Imutável"],
    ["[TODAS]", "Auditoria", "updated_at", "DateTime", "SIM", "SIM", "Data/hora da última atualização", "2026-03-16T14:00:00Z", "Auto", "Atualizado automaticamente"],
    ["[TODAS]", "Auditoria", "created_by", "String FK", "SIM", "PARCIAL", "Usuário que criou", "user_001", "FK → User", "Obrigatório para auditoria"],
    ["[TODAS]", "Auditoria", "updated_by", "String FK", "SIM", "NÃO", "Usuário da última alteração", "user_002", "FK → User", "Rastrear quem mudou"],
    ["[TODAS]", "Lifecycle", "deleted_at", "DateTime?", "NÃO", "NÃO", "Soft delete - inativação lógica", "null ou 2026-04-01T...", "—", "Nunca deletar fisicamente"],
    ["[TODAS]", "Lifecycle", "status", "Enum", "SIM", "SIM", "Estado atual do registro", "ACTIVE / INACTIVE / ARCHIVED", "Enum", "Máquina de estados"],
    ["[TODAS]", "Versionamento", "version", "Integer", "SIM", "NÃO", "Versão do registro (optimistic lock)", "1, 2, 3...", "Auto-incremento", "Previne conflito de edição"],

    # ─── Datas obrigatórias em entidades financeiras ───
    ["OfficialEntry", "Datas", "issue_date", "Date", "SIM", "SIM", "Data de emissão do documento", "2026-03-01", "Date válido", "Quando o título/NF foi emitido"],
    ["OfficialEntry", "Datas", "due_date", "Date", "SIM", "SIM", "Data de vencimento", "2026-03-15", "≥ issue_date", "Quando deve ser pago"],
    ["OfficialEntry", "Datas", "competence_date", "Date", "SIM", "NÃO", "Data de competência (DRE/gerencial)", "2026-03-01", "Date válido", "CAMPO MAIS IMPORTANTE PARA RELATÓRIOS"],
    ["Settlement", "Datas", "payment_date", "Date", "SIM", "SIM", "Data do pagamento efetivo", "2026-03-14", "Date válido", "Quando saiu/entrou dinheiro"],
    ["Settlement", "Datas", "settlement_date", "Date", "SIM", "NÃO", "Data de liquidação formal", "2026-03-14", "≥ payment_date", "Pode diferir do pagamento"],
    ["BankStatementLine", "Datas", "bank_posted_date", "Date", "SIM", "PARCIAL", "Data que o banco registrou", "2026-03-15", "Date válido", "Referência para conciliação"],
    ["StagingEntry", "Datas", "imported_at", "DateTime", "SIM", "SIM", "Quando o dado entrou no sistema", "2026-03-10T08:00:00Z", "Auto", "Rastreabilidade de importação"],

    # ─── Valores obrigatórios ───
    ["OfficialEntry", "Valores", "gross_amount", "Decimal(15,2)", "SIM", "PARCIAL", "Valor bruto original do título", "1500.00", "> 0", "Valor antes de qualquer desconto/acréscimo"],
    ["OfficialEntry", "Valores", "net_amount", "Decimal(15,2)", "SIM", "NÃO", "Valor líquido (gross - descontos + acréscimos)", "1485.00", "> 0", "Valor efetivo esperado"],
    ["Settlement", "Valores", "paid_amount", "Decimal(15,2)", "SIM", "PARCIAL", "Valor efetivamente pago", "1485.00", "> 0", "Pode diferir do net_amount"],
    ["Settlement", "Valores", "discount_amount", "Decimal(15,2)", "NÃO", "NÃO", "Desconto obtido/concedido", "15.00", "≥ 0", "Desconto por antecipação etc."],
    ["Settlement", "Valores", "fine_amount", "Decimal(15,2)", "NÃO", "NÃO", "Multa por atraso", "0.00", "≥ 0", "Multa contratual"],
    ["Settlement", "Valores", "interest_amount", "Decimal(15,2)", "NÃO", "NÃO", "Juros cobrados/pagos", "0.00", "≥ 0", "Juros de mora"],
    ["Settlement", "Valores", "fee_amount", "Decimal(15,2)", "NÃO", "NÃO", "Taxas bancárias/operacionais", "5.50", "≥ 0", "TED, DOC, boleto etc."],
    ["Settlement", "Valores", "balance_remaining", "Decimal(15,2)", "SIM", "NÃO", "Saldo restante após esta baixa", "0.00", "≥ 0", "0 = totalmente liquidado"],

    # ─── Campos de conciliação ───
    ["ReconciliationMatch", "Matching", "match_confidence", "Integer", "SIM", "NÃO", "Score de confiança do match (0-100)", "95", "0-100", "Abaixo de 80 → revisão humana"],
    ["ReconciliationMatch", "Matching", "match_rule_id", "String FK", "SIM", "NÃO", "ID da regra que gerou o match", "rule_exact_value", "FK → MatchRule", "Rastrear qual lógica decidiu"],
    ["ReconciliationMatch", "Matching", "match_basis", "Enum", "SIM", "NÃO", "Base da decisão de match", "VALUE+DATE+DOC", "Enum", "valor/data/documento/descrição/banco"],
    ["ReconciliationMatch", "Matching", "requires_human_review", "Boolean", "SIM", "NÃO", "Precisa de revisão manual?", "false", "Bool", "true se confidence < threshold"],
    ["ReconciliationMatch", "Matching", "review_reason", "String", "NÃO", "NÃO", "Motivo da revisão humana", "Valor diverge 2.5% acima da tolerância", "—", "Justificativa estruturada"],
    ["ReconciliationMatch", "Aprovação", "approved_by", "String FK", "NÃO", "NÃO", "Quem aprovou o match", "user_003", "FK → User", "Null se auto-aprovado"],
    ["ReconciliationMatch", "Aprovação", "approved_at", "DateTime", "NÃO", "NÃO", "Quando foi aprovado", "2026-03-16T15:00:00Z", "—", "Timestamp de aprovação"],

    # ─── Taxonomia (4 camadas) ───
    ["OfficialEntry", "Classificação", "movement_type", "Enum", "SIM", "NÃO", "Tipo de movimento", "ENTRADA / SAIDA / TRANSFERENCIA / AJUSTE", "Enum 4 valores", "Camada 1 – Tipo"],
    ["OfficialEntry", "Classificação", "financial_nature", "Enum", "SIM", "NÃO", "Natureza financeira", "OPERACIONAL / NAO_OPERACIONAL / FINANCEIRA / PATRIMONIAL", "Enum 4 valores", "Camada 2 – Natureza"],
    ["OfficialEntry", "Classificação", "management_category_id", "String FK", "SIM", "SIM", "Categoria gerencial", "cat_materia_prima", "FK → Category", "Camada 3 – Categoria"],
    ["OfficialEntry", "Classificação", "classification_status", "Enum", "SIM", "NÃO", "Status da classificação", "CLASSIFIED / PENDING / DOUBT / CONFLICT", "Enum 4 valores", "Camada 4 – Status (substitui 'DUVIDA' como categoria)"],

    # ─── ImportBatch enriquecido ───
    ["ImportBatch", "Metadados", "file_hash", "String", "SIM", "NÃO", "Hash SHA-256 do arquivo importado", "a3f2b8c9d1...", "SHA-256", "Prova de integridade"],
    ["ImportBatch", "Metadados", "file_size", "Integer", "NÃO", "NÃO", "Tamanho do arquivo em bytes", "245760", "> 0", "Para detectar truncamento"],
    ["ImportBatch", "Metadados", "record_count", "Integer", "SIM", "NÃO", "Qtd de registros no arquivo", "150", "> 0", "Conferir com qtd importada"],
    ["ImportBatch", "Metadados", "total_amount", "Decimal(15,2)", "SIM", "NÃO", "Soma financeira dos registros", "458293.50", "—", "Conferir com soma importada"],
    ["ImportBatch", "Metadados", "parse_errors", "JSON", "NÃO", "NÃO", "Lista de erros de parsing", "[{line:5, error:'date invalid'}]", "JSON array", "Para diagnóstico"],
    ["ImportBatch", "Metadados", "parser_version", "String", "SIM", "NÃO", "Versão do parser usado", "v2.1.0", "SemVer", "Rastreabilidade de processamento"],

    # ─── Parcela enriquecida ───
    ["Installment", "Parcelamento", "parent_entry_id", "String FK", "SIM", "PARCIAL", "Lançamento mãe que gerou a parcela", "entry_abc123", "FK → OfficialEntry", "Rastreabilidade de origem"],
    ["Installment", "Parcelamento", "installment_number", "Integer", "SIM", "PARCIAL", "Número da parcela", "3", "1..total", "Ex: 3 de 12"],
    ["Installment", "Parcelamento", "total_installments", "Integer", "SIM", "PARCIAL", "Total de parcelas do parcelamento", "12", "> 0", "Permite calcular progresso"],
    ["Installment", "Parcelamento", "generation_rule_snapshot", "JSON", "SIM", "NÃO", "Snapshot imutável da regra de geração", "{freq:'monthly', day:15}", "JSON", "Nunca alterar após geração"],
    ["Installment", "Parcelamento", "expected_date", "Date", "SIM", "PARCIAL", "Data prevista de vencimento", "2026-06-15", "Date", "Calculada pela regra"],
    ["Installment", "Parcelamento", "manually_edited", "Boolean", "SIM", "NÃO", "Foi editada manualmente?", "false", "Bool", "Flag de intervenção humana"],
    ["Installment", "Parcelamento", "edited_by", "String FK", "NÃO", "NÃO", "Quem editou", "user_002", "FK → User", "Só se manually_edited=true"],
    ["Installment", "Parcelamento", "edit_reason", "String", "NÃO", "NÃO", "Motivo da edição manual", "Cliente renegociou prazo", "—", "Obrigatório se edited"],

    # ─── Transferência interna ───
    ["InternalTransfer", "Tesouraria", "id", "UUID", "SIM", "NÃO (NOVA)", "ID da transferência", "transfer_001", "Auto", "NOVA ENTIDADE"],
    ["InternalTransfer", "Tesouraria", "tenant_id", "String FK", "SIM", "NÃO (NOVA)", "Empresa", "tenant_abc", "FK → Tenant", "Multi-tenant"],
    ["InternalTransfer", "Tesouraria", "source_account_id", "String FK", "SIM", "NÃO (NOVA)", "Conta de origem", "acct_001", "FK → BankAccount", "De onde saiu"],
    ["InternalTransfer", "Tesouraria", "target_account_id", "String FK", "SIM", "NÃO (NOVA)", "Conta de destino", "acct_002", "FK → BankAccount", "Para onde foi"],
    ["InternalTransfer", "Tesouraria", "amount", "Decimal(15,2)", "SIM", "NÃO (NOVA)", "Valor transferido", "50000.00", "> 0", "Valor único (não duplica)"],
    ["InternalTransfer", "Tesouraria", "transfer_date", "Date", "SIM", "NÃO (NOVA)", "Data da transferência", "2026-03-15", "Date", "Referência"],
    ["InternalTransfer", "Tesouraria", "debit_entry_id", "String FK", "SIM", "NÃO (NOVA)", "Lançamento de saída", "entry_debit_001", "FK → OfficialEntry", "Par de saída"],
    ["InternalTransfer", "Tesouraria", "credit_entry_id", "String FK", "SIM", "NÃO (NOVA)", "Lançamento de entrada", "entry_credit_001", "FK → OfficialEntry", "Par de entrada"],
    ["InternalTransfer", "Tesouraria", "reference", "String", "NÃO", "NÃO (NOVA)", "Referência/descrição", "Cobertura de caixa filial SP", "—", "Livre"],
]

add_sheet("20-Modelo Dados Alvo", mdl_headers, mdl_data,
          highlight_col=6,
          highlight_vals={"alert": ["NÃO", "NÃO (NOVA)"], "warn": ["PARCIAL"], "ok": ["SIM"]})

print("✅ Aba 20 criada")

# ═══════════════════════════════════════════════════════════════════════
# ABA 21 – MOTOR DE REGRAS
# ═══════════════════════════════════════════════════════════════════════
rules_headers = [
    "ID Regra", "Categoria", "Nome", "Condição (SE...)",
    "Ação (ENTÃO...)", "Prioridade", "Tipo de Ação",
    "Score/Confiança", "Exceção se Falhar",
    "Módulo Afetado", "Exemplo Prático", "Status Atual"
]
rules_data = [
    ["MR01", "Classificação", "Auto-classificar por descrição extrato",
     "SE origem = extrato_bancario E descrição CONTÉM palavra-chave da ClassificationRule",
     "ENTÃO sugerir categoria da regra com score baseado em hits anteriores",
     "1 (Alta)", "Sugestão", "70-100 conforme histórico",
     "Encaminhar para fila de revisão manual",
     "Classificação", "Descrição 'PGTO NF 1234 FORNECEDOR X' → categoria MATERIA_PRIMA (score 92)",
     "PARCIAL – ClassificationRule existe mas sem score"],

    ["MR02", "Classificação", "Auto-classificar por fornecedor conhecido",
     "SE fornecedor_id já tem ≥3 lançamentos anteriores com mesma categoria",
     "ENTÃO aplicar categoria mais frequente com score baseado em % de ocorrência",
     "2", "Sugestão", "60-95 conforme consistência",
     "Encaminhar para fila se score < 80",
     "Classificação", "Fornecedor 'Gráfica ABC' → 100% das vezes é MARKETING → score 95",
     "NÃO EXISTE"],

    ["MR03", "Classificação", "Bloquear classificação incoerente",
     "SE tipo_movimento = SAIDA E categoria = RECEITA",
     "ENTÃO bloquear incorporação, marcar como CONFLITO",
     "0 (Máxima)", "Bloqueio", "N/A",
     "Impedir avanço no pipeline até resolução",
     "Classificação / Staging", "Saída de R$ 5.000 classificada como RECEITA → BLOQUEADO",
     "NÃO EXISTE"],

    ["MR04", "Conciliação", "Match exato por valor + data",
     "SE settlement.paid_amount = bankline.amount E payment_date = bank_posted_date",
     "ENTÃO criar ReconciliationMatch com confidence=100, basis=VALUE+DATE",
     "1 (Alta)", "Auto-match", "100",
     "N/A – match perfeito",
     "Conciliação", "Baixa R$ 1.500 em 15/03 = Extrato R$ 1.500 em 15/03 → match 100%",
     "PARCIAL – match existe mas sem confidence"],

    ["MR05", "Conciliação", "Match por valor com tolerância",
     "SE |settlement.paid_amount - bankline.amount| ≤ tolerância (ex: R$ 0.50 ou 0.1%)",
     "ENTÃO criar match com confidence=85, basis=VALUE_APPROX, requires_human_review=true",
     "2", "Match + Revisão", "80-90",
     "Encaminhar para fila de revisão com motivo 'valor divergente'",
     "Conciliação", "Baixa R$ 1.500 vs Extrato R$ 1.500.30 (taxa bancária) → match 85%, revisar",
     "NÃO EXISTE"],

    ["MR06", "Conciliação", "Detectar match múltiplo (N:1 ou 1:N)",
     "SE soma de N settlements = bankline.amount (ou vice-versa)",
     "ENTÃO criar grupo de match com confidence=75, basis=AGGREGATE, requires_human_review=true",
     "3", "Match Grupo + Revisão", "70-80",
     "Encaminhar para revisão com todos os candidatos listados",
     "Conciliação", "3 pagamentos (R$ 500 + R$ 300 + R$ 200) = 1 linha extrato R$ 1.000",
     "NÃO EXISTE"],

    ["MR07", "Conciliação", "Bloquear auto-match acima de tolerância",
     "SE divergência > tolerância máxima (ex: 5% ou R$ 100)",
     "ENTÃO NÃO criar match, encaminhar para fila com motivo 'divergência excessiva'",
     "0 (Máxima)", "Bloqueio", "N/A",
     "Registro fica em 'pendente de conciliação'",
     "Conciliação", "Baixa R$ 10.000 vs Extrato R$ 9.500 (5% diff) → SEM match, revisar",
     "NÃO EXISTE"],

    ["MR08", "Período/Corte", "Revisão de corte temporal",
     "SE payment_date fora do período MAS bank_posted_date dentro do período",
     "ENTÃO marcar como 'revisão de corte', não excluir automaticamente",
     "1 (Alta)", "Flag + Revisão", "N/A",
     "Incluir na conciliação com flag de corte para revisão CFO",
     "Conciliação / Relatórios", "Pagamento 28/02 compensado 01/03 → incluir em março com flag",
     "NÃO EXISTE – causa do bug das 868 linhas"],

    ["MR09", "Validação", "Obrigar campos mínimos para incorporação",
     "SE StagingEntry não tem: tenant_id, competence_date, gross_amount, movement_type, category",
     "ENTÃO bloquear transição para OfficialEntry, status = 'VALIDACAO_PENDENTE'",
     "0 (Máxima)", "Bloqueio", "N/A",
     "Listar campos faltantes na mensagem de erro",
     "Staging → Oficial", "Entrada sem competence_date → NÃO incorporar, pedir preenchimento",
     "PARCIAL – validação existe mas sem competence_date"],

    ["MR10", "Validação", "Detectar duplicidade na importação",
     "SE novo registro tem mesmo (tenant_id + source_system + source_record_id + amount + date)",
     "ENTÃO bloquear como DUPLICADO, não criar StagingEntry",
     "0 (Máxima)", "Bloqueio", "N/A",
     "Mostrar registro existente para comparação",
     "Importação", "Mesmo arquivo OFX importado 2 vezes → segunda vez bloqueia",
     "PARCIAL – verificação existe mas critério simples"],

    ["MR11", "Validação", "Bloquear edição de parcela sem justificativa",
     "SE Installment.manually_edited = true E edit_reason está vazio",
     "ENTÃO rejeitar alteração, exigir motivo",
     "1 (Alta)", "Bloqueio", "N/A",
     "Campo edit_reason torna-se obrigatório",
     "Parcelamento", "Alterar data da parcela 3/12 → obrigar justificativa 'renegociação'",
     "NÃO EXISTE"],

    ["MR12", "Transferência", "Auto-parear transferências internas",
     "SE dois lançamentos na mesma data, mesmo valor, contas diferentes, um ENTRADA e outro SAIDA, mesma empresa",
     "ENTÃO criar InternalTransfer vinculando os dois, excluir de DRE",
     "2", "Auto-match + Exclusão DRE", "85-95",
     "Se não encontrar par → fila 'transferência pendente de pareamento'",
     "Tesouraria", "Saída R$ 50k conta Itaú + Entrada R$ 50k conta BB mesmo dia → transferência",
     "NÃO EXISTE"],

    ["MR13", "Exceção", "Fila de exceções com motivo estruturado",
     "SE qualquer regra acima gerar exceção/bloqueio",
     "ENTÃO criar registro em ExceptionQueue com: regra_id, motivo, dados envolvidos, sugestão, prioridade",
     "N/A", "Encaminhamento", "N/A",
     "Itens ficam na fila até resolução manual ou nova regra",
     "Todos", "Qualquer bloqueio → aparece na inbox operacional do analista com contexto completo",
     "NÃO EXISTE"],

    ["MR14", "Auditoria", "Log obrigatório em decisão automática",
     "SE qualquer regra MR01-MR13 for executada",
     "ENTÃO criar AuditLog com: rule_id, input_data, output_decision, confidence, timestamp",
     "N/A", "Log", "N/A",
     "Se log falhar, bloquear a ação (fail-safe)",
     "Auditoria", "MR04 executou match → AuditLog: 'rule=MR04, confidence=100, matched entry_123 ↔ line_456'",
     "NÃO EXISTE"],
]

add_sheet("21-Motor de Regras", rules_headers, rules_data,
          highlight_col=12,
          highlight_vals={"alert": ["NÃO EXISTE"], "warn": ["PARCIAL – ClassificationRule existe mas sem score", "PARCIAL – match existe mas sem confidence", "PARCIAL – validação existe mas sem competence_date", "PARCIAL – verificação existe mas critério simples"]})

print("✅ Aba 21 criada")

# ═══════════════════════════════════════════════════════════════════════
# ATUALIZAR ABA 13 – PROBLEMAS/RISCOS (adicionar novos)
# ═══════════════════════════════════════════════════════════════════════
ws13 = wb["13-Problemas Riscos"]
# Descobrir última linha com dado
max_row_13 = ws13.max_row
max_col_13 = ws13.max_column

new_problems = [
    ["P21", "Ambiguidade de datas causa exclusão incorreta de registros",
     "CRITICO", "IMEDIATO", "Conciliação / Relatórios",
     "O sistema usa data de emissão (com horário) em vez de data de pagamento normalizada para filtrar períodos. 868 registros (R$ 267.724,53) foram excluídos incorretamente da conciliação.",
     "Cada processo deve declarar qual data usa. Implementar competence_date como campo distinto.", "SIM",
     "Conciliação CFO jan-fev: parecer 'não aprovar sem ajuste'",
     "RA01", "Contamina DRE, aging, overdue e caixa", "Todas as entidades financeiras"],

    ["P22", "Staging sem máquina de estados formal",
     "CRITICO", "IMEDIATO", "Pipeline de Dados",
     "StagingEntry tem campo status mas sem transições controladas. Dados podem pular etapas e chegar ao OfficialEntry sem validação completa.",
     "Implementar state machine com 9 estados formais e transições auditadas.", "NÃO",
     "Análise de código: transições não são validadas",
     "RA02", "Dados corrompidos chegam ao oficial sem filtro", "StagingEntry"],

    ["P23", "Categorias misturam semânticas diferentes",
     "ALTO", "CURTO PRAZO", "Classificação",
     "RECEITA, DUVIDA, SALARIO, TRANSFERENCIA ENTRE CONTAS, BEBIDAS coexistem como categorias. 'DUVIDA' é status, não categoria. 'TRANSFERENCIA' é tipo de movimento, não natureza.",
     "Separar em 4 camadas: tipo_movimento, natureza_financeira, categoria_gerencial, status_classificação.", "NÃO",
     "Base de dados: campo category mistura semânticas",
     "RA05", "Relatórios gerenciais incoerentes", "OfficialEntry, Category"],

    ["P24", "Transferências inflam receitas e despesas",
     "ALTO", "CURTO PRAZO", "Tesouraria",
     "Transferências entre contas da mesma empresa são tratadas como entrada+saída normais, inflando DRE e distorcendo fluxo de caixa.",
     "Criar entidade InternalTransfer com pareamento automático. Excluir de DRE.", "NÃO",
     "Categoria TRANSFERENCIA ENTRE CONTAS sem lógica de pareamento",
     "RA06", "Fluxo de caixa distorcido, DRE inflado", "OfficialEntry, BankAccount"],

    ["P25", "Conciliação sem rastreabilidade de decisão",
     "ALTO", "CURTO PRAZO", "Conciliação",
     "ReconciliationMatch não registra: score de confiança, regra que gerou match, base da decisão, se precisa revisão humana, quem aprovou.",
     "Adicionar campos de auditoria: match_confidence, match_rule_id, match_basis, requires_human_review, approved_by/at.", "NÃO",
     "Schema ReconciliationMatch: campos de auditoria ausentes",
     "RA04", "Impossível auditar matches, melhorar algoritmo", "ReconciliationMatch"],

    ["P26", "ImportBatch sem metadados de integridade",
     "MEDIO", "MEDIO PRAZO", "Importação",
     "Importações não guardam hash do arquivo, contagem de registros, soma financeira, erros de parsing, versão do parser. Impossível provar integridade.",
     "Enriquecer ImportBatch com: file_hash, record_count, total_amount, parse_errors, parser_version.", "NÃO",
     "Schema ImportBatch: metadados limitados",
     "RA09", "Sem prova de origem, erros silenciosos", "ImportBatch"],

    ["P27", "Parcelas sem snapshot da regra de geração",
     "MEDIO", "MEDIO PRAZO", "Parcelamento",
     "Se a regra de recorrência muda, parcelas futuras mudam retroativamente. Sem registro de edição manual e motivo.",
     "Cada geração cria snapshot imutável. Campos: generation_rule_snapshot, manually_edited, edited_by, edit_reason.", "NÃO",
     "RecurringRule + Installment: sem versionamento",
     "RA08", "Parcelas inconsistentes, risco de fraude operacional", "Installment, RecurringRule"],
]

for i, row_data in enumerate(new_problems):
    r = max_row_13 + 1 + i
    for c, val in enumerate(row_data, 1):
        ws13.cell(row=r, column=c, value=val)
    # Style
    for c in range(1, max_col_13 + 1):
        cell = ws13.cell(row=r, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP
    # Highlight critical/high
    sev = ws13.cell(row=r, column=3).value
    if sev == "CRITICO":
        for c in range(1, max_col_13 + 1):
            ws13.cell(row=r, column=c).fill = ALERT_FILL
    elif sev == "ALTO":
        for c in range(1, max_col_13 + 1):
            ws13.cell(row=r, column=c).fill = WARN_FILL

# Update auto-filter
ws13.auto_filter.ref = f"A1:{get_column_letter(max_col_13)}{max_row_13 + len(new_problems)}"

print(f"✅ Aba 13 atualizada: +{len(new_problems)} problemas (P21-P27)")

# ═══════════════════════════════════════════════════════════════════════
# ATUALIZAR ABA 14 – BACKLOG (adicionar novas melhorias)
# ═══════════════════════════════════════════════════════════════════════
ws14 = wb["14-Backlog Melhorias"]
max_row_14 = ws14.max_row
max_col_14 = ws14.max_column

new_backlog = [
    ["M21", "Implementar regra mestra de datas e períodos",
     "P0", "CRÍTICA", "Conciliação / Relatórios",
     "Definir explicitamente qual campo de data governa cada processo. Implementar competence_date, bank_posted_date, settlement_date como campos distintos obrigatórios.",
     "Resolve P21 (868 linhas excluídas). Fundação para toda a conciliação.", "2-3 sprints",
     "Refatoração de queries + migração + testes extensivos", "RA01, P21",
     "Sistema inteiro", "Datas corretas = números corretos", "PLANEJAMENTO"],

    ["M22", "Implementar state machine no staging",
     "P0", "CRÍTICA", "Pipeline de Dados",
     "Criar máquina de estados formal com 9 estados e transições auditadas. Bloquear incorporação sem passar por todos os estados obrigatórios.",
     "Resolve P22. Garante qualidade dos dados antes de virar oficial.", "1-2 sprints",
     "Enum de estados + validação de transições + AuditLog", "RA02, P22",
     "StagingEntry, AuditLog", "Quarentena real, não só conceitual", "PLANEJAMENTO"],

    ["M23", "Construir motor de regras de classificação",
     "P1", "ALTA", "Classificação",
     "Motor com condições compostas (AND/OR), ações (classificar/bloquear/alertar), prioridade, score de confiança, log de regra aplicada.",
     "Resolve P23 parcialmente. Elimina dependência de operador experiente.", "3-4 sprints",
     "Novo módulo: rules engine + admin UI", "RA03",
     "ClassificationRule, StagingEntry", "Automação real com auditoria", "BACKLOG"],

    ["M24", "Adicionar score de confiança na conciliação",
     "P1", "ALTA", "Conciliação",
     "Campos: match_confidence, match_rule_id, match_basis, requires_human_review, review_reason, approved_by, approved_at. Fila de revisão para matches abaixo do threshold.",
     "Resolve P25. Conciliação auditável.", "2 sprints",
     "Novos campos + lógica de fila + UI de revisão", "RA04, P25",
     "ReconciliationMatch", "Transformar conciliação em processo, não palpite", "BACKLOG"],

    ["M25", "Reestruturar taxonomia em 4 camadas",
     "P1", "ALTA", "Dados Mestres",
     "Separar: tipo_movimento (4 valores), natureza_financeira (4 valores), categoria_gerencial (FK), status_classificação (4 valores). Migrar dados existentes.",
     "Resolve P23, P24 parcialmente. Relatórios confiáveis.", "2-3 sprints",
     "Reestruturação de dados mestres + migração + UI", "RA05, P23",
     "OfficialEntry, Category", "Cada campo tem semântica única e estável", "BACKLOG"],

    ["M26", "Modelar transferências internas como entidade própria",
     "P1", "ALTA", "Tesouraria",
     "Criar InternalTransfer: conta_origem, conta_destino, valor, data, referência, debit_entry_id, credit_entry_id. Auto-pareamento. Exclusão de DRE.",
     "Resolve P24. Fluxo de caixa e DRE corretos.", "1-2 sprints",
     "Nova entidade + lógica de pareamento + exclusão DRE", "RA06, P24",
     "OfficialEntry, BankAccount, Settlement", "Transferência ≠ receita nem despesa", "BACKLOG"],

    ["M27", "Tornar OfficialEntry imutável após incorporação",
     "P2", "MÉDIA", "Financeiro",
     "Após status=INCORPORATED, bloquear edição de campos financeiros. Toda alteração via Settlement. Preservar histórico original.",
     "Resolve RA07. Auditoria de títulos garantida.", "1 sprint",
     "Validação no update + Settlement para alterações", "RA07",
     "OfficialEntry, Settlement, AuditLog", "Título original sempre intacto", "BACKLOG"],

    ["M28", "Implementar versionamento de parcelas/recorrências",
     "P2", "MÉDIA", "Parcelamento",
     "Snapshot imutável da regra na geração. Campos: generation_rule_snapshot, manually_edited, edited_by, edit_reason.",
     "Resolve P27. Rastreabilidade total.", "1 sprint",
     "Campos novos + snapshot logic + UI de histórico", "RA08, P27",
     "Installment, RecurringRule", "Cada parcela sabe por que existe", "BACKLOG"],

    ["M29", "Enriquecer ImportBatch com metadados de integridade",
     "P2", "MÉDIA", "Importação",
     "Campos: file_hash (SHA-256), file_size, record_count, total_amount, parse_errors[], parser_version, source_type, raw_file_stored.",
     "Resolve P26. Prova de origem e integridade.", "1 sprint",
     "Campos novos na entidade existente", "RA09, P26",
     "ImportBatch, StagingEntry", "Provar de onde veio cada dado", "BACKLOG"],

    ["M30", "Criar fila de exceções operacional",
     "P1", "ALTA", "Operações",
     "Entidade ExceptionQueue: regra_id, motivo estruturado, dados envolvidos, sugestão, prioridade. Inbox para analistas. Dashboard de exceções.",
     "Complementa MR13. Centro nervoso operacional.", "2 sprints",
     "Nova entidade + UI de inbox + dashboard", "RA03, MR13",
     "Todos os módulos", "Tudo que não bater em regra tem destino claro", "BACKLOG"],
]

for i, row_data in enumerate(new_backlog):
    r = max_row_14 + 1 + i
    for c, val in enumerate(row_data, 1):
        ws14.cell(row=r, column=c, value=val)
    for c in range(1, max_col_14 + 1):
        cell = ws14.cell(row=r, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP
    prio = ws14.cell(row=r, column=3).value
    if prio == "P0":
        for c in range(1, max_col_14 + 1):
            ws14.cell(row=r, column=c).fill = ALERT_FILL
    elif prio == "P1":
        for c in range(1, max_col_14 + 1):
            ws14.cell(row=r, column=c).fill = WARN_FILL

ws14.auto_filter.ref = f"A1:{get_column_letter(max_col_14)}{max_row_14 + len(new_backlog)}"

print(f"✅ Aba 14 atualizada: +{len(new_backlog)} melhorias (M21-M30)")

# ═══════════════════════════════════════════════════════════════════════
# ATUALIZAR ABA 02 – DICIONÁRIO DE DADOS (adicionar campos faltantes)
# ═══════════════════════════════════════════════════════════════════════
ws02 = wb["02-Dicionario de Dados"]
max_row_02 = ws02.max_row
max_col_02 = ws02.max_column

new_dict_entries = [
    # Datas que faltam
    ["OfficialEntry", "competence_date", "Date", "—", "SIM",
     "Data de competência para DRE e relatórios gerenciais",
     "2026-03-01", "Formato YYYY-MM-DD",
     "CAMPO MAIS IMPORTANTE – governa relatórios, DRE, período contábil. Não confundir com issue_date ou payment_date.",
     "NÃO EXISTE", "RA01", "RN-NOVA", "P0"],
    ["Settlement", "settlement_date", "Date", "—", "SIM",
     "Data de liquidação formal (pode diferir de payment_date)",
     "2026-03-14", "≥ payment_date",
     "Data que o banco formalmente reconhece a liquidação.",
     "NÃO EXISTE", "RA01", "RN-NOVA", "P0"],
    ["BankStatementLine", "bank_posted_date", "Date", "—", "SIM",
     "Data que o banco registrou a transação no extrato",
     "2026-03-15", "Date válido",
     "Referência primária para conciliação bancária.",
     "PARCIAL", "RA01", "RN-NOVA", "P0"],

    # Valores que faltam
    ["OfficialEntry", "gross_amount", "Decimal(15,2)", "—", "SIM",
     "Valor bruto original do título, antes de qualquer desconto ou acréscimo",
     "1500.00", "> 0",
     "Valor de face do documento. Nunca alterar após incorporação.",
     "PARCIAL (amount)", "RA07", "RN-NOVA", "P1"],
    ["OfficialEntry", "net_amount", "Decimal(15,2)", "—", "SIM",
     "Valor líquido esperado (gross - descontos + acréscimos)",
     "1485.00", "> 0",
     "Valor real esperado de pagamento/recebimento.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P1"],
    ["Settlement", "discount_amount", "Decimal(15,2)", "—", "NÃO",
     "Desconto obtido ou concedido na baixa",
     "15.00", "≥ 0",
     "Desconto por antecipação, bonificação etc.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P2"],
    ["Settlement", "fine_amount", "Decimal(15,2)", "—", "NÃO",
     "Multa por atraso aplicada na baixa",
     "0.00", "≥ 0",
     "Multa contratual, geralmente 2%.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P2"],
    ["Settlement", "interest_amount", "Decimal(15,2)", "—", "NÃO",
     "Juros de mora cobrados ou pagos",
     "12.50", "≥ 0",
     "Juros pro-rata por dias de atraso.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P2"],
    ["Settlement", "fee_amount", "Decimal(15,2)", "—", "NÃO",
     "Taxas bancárias/operacionais",
     "5.50", "≥ 0",
     "TED, DOC, boleto, cartão etc.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P2"],
    ["Settlement", "balance_remaining", "Decimal(15,2)", "—", "SIM",
     "Saldo restante do título após esta baixa",
     "0.00", "≥ 0",
     "0.00 = totalmente liquidado. >0 = baixa parcial.",
     "NÃO EXISTE", "RA07", "RN-NOVA", "P1"],

    # Campos de conciliação
    ["ReconciliationMatch", "match_confidence", "Integer (0-100)", "—", "SIM",
     "Score de confiança do match automático",
     "95", "0-100",
     "Abaixo de 80 → revisão humana obrigatória.",
     "NÃO EXISTE", "RA04", "MR04-MR07", "P1"],
    ["ReconciliationMatch", "match_rule_id", "String FK", "—", "SIM",
     "ID da regra que gerou o match",
     "MR04", "FK → MatchRule",
     "Rastrear qual lógica decidiu. Essencial para melhorar o algoritmo.",
     "NÃO EXISTE", "RA04", "MR04-MR07", "P1"],
    ["ReconciliationMatch", "match_basis", "Enum", "—", "SIM",
     "Base da decisão: quais critérios foram usados",
     "VALUE+DATE+DOC", "Enum",
     "Valores possíveis: VALUE, DATE, DOC, DESCRIPTION, BANK, AGGREGATE.",
     "NÃO EXISTE", "RA04", "MR04-MR07", "P1"],
    ["ReconciliationMatch", "requires_human_review", "Boolean", "—", "SIM",
     "Flag indicando se precisa de revisão manual",
     "false", "true/false",
     "true quando confidence < threshold ou regra exige revisão.",
     "NÃO EXISTE", "RA04", "MR05-MR07", "P1"],
    ["ReconciliationMatch", "approved_by", "String FK", "—", "NÃO",
     "Usuário que aprovou o match (se revisado manualmente)",
     "user_003", "FK → User",
     "Null se auto-aprovado (confidence ≥ threshold).",
     "NÃO EXISTE", "RA04", "—", "P1"],
    ["ReconciliationMatch", "approved_at", "DateTime", "—", "NÃO",
     "Timestamp de aprovação do match",
     "2026-03-16T15:00:00Z", "DateTime",
     "Complementa approved_by para trilha de auditoria completa.",
     "NÃO EXISTE", "RA04", "—", "P1"],

    # Campos de classificação (4 camadas)
    ["OfficialEntry", "movement_type", "Enum", "—", "SIM",
     "Tipo de movimento: ENTRADA, SAIDA, TRANSFERENCIA, AJUSTE",
     "SAIDA", "Enum 4 valores",
     "Camada 1 da taxonomia. Substitui a mistura atual de categorias.",
     "NÃO EXISTE", "RA05", "MR03", "P1"],
    ["OfficialEntry", "financial_nature", "Enum", "—", "SIM",
     "Natureza financeira: OPERACIONAL, NAO_OPERACIONAL, FINANCEIRA, PATRIMONIAL",
     "OPERACIONAL", "Enum 4 valores",
     "Camada 2 da taxonomia. Governa classificação contábil.",
     "NÃO EXISTE", "RA05", "MR03", "P1"],
    ["OfficialEntry", "classification_status", "Enum", "—", "SIM",
     "Status da classificação: CLASSIFIED, PENDING, DOUBT, CONFLICT",
     "CLASSIFIED", "Enum 4 valores",
     "Camada 4 da taxonomia. 'DUVIDA' agora é status, não categoria.",
     "NÃO EXISTE", "RA05", "MR01-MR03", "P1"],

    # Campos de auditoria universal
    ["[TODAS]", "updated_by", "String FK", "—", "SIM",
     "Usuário responsável pela última alteração",
     "user_002", "FK → User",
     "Rastreabilidade obrigatória. Quem tocou por último.",
     "NÃO EXISTE", "RA02", "—", "P0"],
    ["[TODAS]", "deleted_at", "DateTime?", "—", "NÃO",
     "Soft delete – inativação lógica sem exclusão física",
     "null", "DateTime ou null",
     "Nunca deletar fisicamente registros financeiros.",
     "NÃO EXISTE", "—", "—", "P2"],
    ["[TODAS]", "version", "Integer", "—", "SIM",
     "Versão do registro para controle de concorrência (optimistic locking)",
     "1", "Auto-incremento",
     "Previne conflito de edição simultânea. Incrementa a cada save.",
     "NÃO EXISTE", "—", "—", "P2"],
]

for i, row_data in enumerate(new_dict_entries):
    r = max_row_02 + 1 + i
    for c, val in enumerate(row_data, 1):
        ws02.cell(row=r, column=c, value=val)
    for c in range(1, max_col_02 + 1):
        cell = ws02.cell(row=r, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP
    exists_val = ws02.cell(row=r, column=10).value
    if exists_val and "NÃO EXISTE" in str(exists_val):
        for c in range(1, max_col_02 + 1):
            ws02.cell(row=r, column=c).fill = ALERT_FILL
    elif exists_val and "PARCIAL" in str(exists_val):
        for c in range(1, max_col_02 + 1):
            ws02.cell(row=r, column=c).fill = WARN_FILL

ws02.auto_filter.ref = f"A1:{get_column_letter(max_col_02)}{max_row_02 + len(new_dict_entries)}"

print(f"✅ Aba 02 atualizada: +{len(new_dict_entries)} campos novos no dicionário")

# ═══════════════════════════════════════════════════════════════════════
# REORDENAR ABAS (garantir ordem correta 00..21)
# ═══════════════════════════════════════════════════════════════════════
desired_order = [
    "00-Guia de Leitura",
    "01-Visao Geral",
    "02-Dicionario de Dados",
    "03-Tabelas Entidades",
    "04-Relacionamentos",
    "05-Regras de Negocio",
    "06-Fluxo da Informacao",
    "07-Modulos Funcionalidades",
    "08-Telas Interfaces",
    "09-Entradas Process Saidas",
    "10-Integracoes",
    "11-Perfis Permissoes",
    "12-Logs Auditoria",
    "13-Problemas Riscos",
    "14-Backlog Melhorias",
    "15-Glossario",
    "16-Padroes Nomenclatura",
    "17-Mudancas Seguras",
    "18-Enums do Sistema",
    "19-Revisao Arquitetural",
    "20-Modelo Dados Alvo",
    "21-Motor de Regras",
]

current_names = wb.sheetnames
new_order = []
for name in desired_order:
    if name in current_names:
        new_order.append(current_names.index(name))

if len(new_order) == len(current_names):
    wb.move_sheet("00-Guia de Leitura", offset=0)
    # Reorder by moving each sheet to correct position
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

print(f"Ordem final: {wb.sheetnames}")

# ═══════════════════════════════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════════════════════════════
wb.save(FILE)
print(f"\n🎉 Excel atualizado com sucesso!")
print(f"   Arquivo: {FILE}")
print(f"   Total abas: {len(wb.sheetnames)}")
print(f"   Novas abas: 19-Revisão Arquitetural, 20-Modelo Dados Alvo, 21-Motor de Regras")
print(f"   Abas atualizadas: 02-Dicionário (+{len(new_dict_entries)} campos), 13-Problemas (+{len(new_problems)}), 14-Backlog (+{len(new_backlog)})")
