#!/usr/bin/env python3
"""Mapa Mestre - Parte 2: Abas 2 (Dicionario), 7-12"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

WB_PATH = os.path.join(os.path.dirname(__file__), "Mapa_Mestre_Sistema_Financeiro.xlsx")
wb = openpyxl.load_workbook(WB_PATH)

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)

def auto_width(ws, num_cols, max_w=40):
    for col in range(1, num_cols + 1):
        mx = 10
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(col)].width = mx + 2

def add_sheet(wb, title, headers, data):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    style_header(ws, len(headers))
    for row_data in data:
        ws.append(row_data)
    if data:
        style_data(ws, 2, len(data) + 1, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))
    return ws

# =====================================================
# ABA 2: DICIONARIO DE DADOS (campos principais)
# =====================================================
headers = ["Campo Tecnico", "Nome Amigavel", "Descricao", "Modulo", "Tabela", "Tipo Dado", "Obrigatorio", "Valor Padrao", "Exemplo", "Origem", "Destino/Uso", "Impacto", "Obs"]
dd = [
    # Tenant
    ["id", "ID", "Identificador unico (CUID)", "Admin", "Tenant", "String", "Sim", "cuid()", "cm1abc2de3f", "Gerado automatico", "FK em todas tabelas", "Critico - chave de isolamento", ""],
    ["name", "Nome da Empresa", "Razao social da empresa", "Admin", "Tenant", "String", "Sim", "-", "Empresa ABC Ltda", "Cadastro manual", "Header, relatorios, listagens", "Alto", ""],
    ["cnpj", "CNPJ", "Cadastro Nacional PJ", "Admin", "Tenant", "String", "Sim", "-", "12.345.678/0001-90", "Cadastro manual", "Identificacao fiscal", "Alto", "Sem validacao de digito"],
    ["slug", "Slug", "Identificador URL-friendly unico", "Admin", "Tenant", "String", "Sim (UNIQUE)", "-", "empresa-abc", "Gerado do nome", "URLs e rotas", "Medio", ""],
    ["active", "Ativo", "Se a empresa esta ativa", "Admin", "Tenant", "Boolean", "Sim", "true", "true", "Cadastro", "Filtro de acesso", "Alto - empresa inativa bloqueia acesso", "Soft delete"],
    # OfficialEntry (mais importante)
    ["sequentialNumber", "Numero Sequencial", "Numero unico do lancamento na empresa", "Financeiro", "OfficialEntry", "Int", "Sim", "Auto-gerado", "1042", "Incorporacao (max+1)", "Identificacao em tela e relatorios", "Critico - rastreabilidade", "BUG: race condition"],
    ["date", "Data Lancamento", "Data da operacao financeira", "Financeiro", "OfficialEntry", "DateTime", "Sim", "-", "2026-03-15", "Staging ou manual", "Filtros, relatorios, DRE", "Critico", ""],
    ["competenceDate", "Data Competencia", "Data do regime de competencia contabil", "Financeiro", "OfficialEntry", "DateTime", "Sim", "-", "2026-03-01", "Incorporacao", "DRE, relatorios contabeis", "Critico", "BUG: forcado = date na incorporacao"],
    ["description", "Descricao", "Descricao do lancamento", "Financeiro", "OfficialEntry", "String", "Sim", "-", "Pagamento fornecedor X - NF 123", "Staging/manual", "Listagens, busca, conciliacao", "Alto", ""],
    ["amount", "Valor", "Valor principal do lancamento", "Financeiro", "OfficialEntry", "Decimal", "Sim", "-", "1500.00", "Staging/manual", "Calculos, saldos, relatorios, KPIs", "Critico", "Armazenado como Decimal do Prisma"],
    ["type", "Tipo Transacao", "Credito ou Debito", "Financeiro", "OfficialEntry", "Enum", "Sim", "-", "DEBIT", "Staging", "Determina direcao do fluxo", "Critico", "CREDIT ou DEBIT"],
    ["status", "Status", "Estado atual do lancamento", "Financeiro", "OfficialEntry", "Enum", "Sim", "OPEN", "SETTLED", "Incorporacao / Liquidacao", "Filtros, relatorios, indicadores", "Critico", "OPEN, PARTIAL, SETTLED, CANCELLED"],
    ["category", "Categoria", "Tipo: Pagar, Receber, Transferencia, Ajuste", "Financeiro", "OfficialEntry", "Enum", "Sim", "-", "PAYABLE", "Derivado do type na incorporacao", "Separa payables de receivables", "Critico", "BUG: derivacao simplista"],
    ["dueDate", "Vencimento", "Data de vencimento do titulo", "Financeiro", "OfficialEntry", "DateTime", "Nao", "-", "2026-04-15", "Staging ou manual", "Calendar, aging, overdue check", "Alto", "BUG: staging nao tem este campo"],
    ["paidAmount", "Valor Pago", "Total ja pago/recebido", "Financeiro", "OfficialEntry", "Decimal", "Nao", "-", "750.00", "Calculado nas liquidacoes", "Saldo restante, status", "Alto", "Soma dos settlements"],
    ["paidDate", "Data Pagamento", "Data da ultima liquidacao total", "Financeiro", "OfficialEntry", "DateTime", "Nao", "-", "2026-03-20", "Liquidacao quando status=SETTLED", "Relatorios de fluxo de caixa", "Medio", "So preenchido quando SETTLED"],
    ["interestAmount", "Juros", "Valor total de juros cobrados", "Financeiro", "OfficialEntry", "Decimal", "Sim", "0", "15.00", "Liquidacao", "Soma acumulada de juros", "Medio", "Nao calculado automaticamente"],
    ["fineAmount", "Multa", "Valor total de multa cobrada", "Financeiro", "OfficialEntry", "Decimal", "Sim", "0", "30.00", "Liquidacao", "Soma acumulada de multas", "Medio", "Nao calculado automaticamente"],
    ["discountAmount", "Desconto", "Valor total de desconto concedido", "Financeiro", "OfficialEntry", "Decimal", "Sim", "0", "10.00", "Liquidacao", "Soma acumulada de descontos", "Medio", ""],
    ["installmentGroupId", "Grupo Parcela", "ID que agrupa parcelas do mesmo titulo", "Financeiro", "OfficialEntry", "String", "Nao", "-", "cm1orig123", "Parcelamento", "Agrupar parcelas na tela", "Medio", "= ID do lancamento original"],
    ["installmentNumber", "Num Parcela", "Numero da parcela (1, 2, 3...)", "Financeiro", "OfficialEntry", "Int", "Nao", "-", "2", "Parcelamento", "Exibicao na tela de parcelas", "Baixo", ""],
    ["stagingEntryId", "ID Staging Origem", "Referencia ao staging que originou", "Financeiro", "OfficialEntry", "String", "Nao (UNIQUE)", "-", "cm1stg456", "Incorporacao", "Rastreabilidade staging->oficial", "Medio", "1:1 com StagingEntry"],
    # StagingEntry
    ["source", "Origem", "De onde veio o lancamento provisorio", "Staging", "StagingEntry", "Enum", "Sim", "-", "IMPORT_BANK_STATEMENT", "Importacao/manual", "Filtros, rastreabilidade", "Medio", "7 opcoes: MANUAL, IMPORT_*"],
    ["counterpartCnpjCpf", "CNPJ/CPF Contraparte", "CNPJ ou CPF do fornecedor/cliente", "Staging", "StagingEntry", "String", "Nao", "-", "12.345.678/0001-90", "Extrato bancario / NF", "Auto-classificacao por CNPJ", "Medio", "Usado na ClassificationRule"],
    ["counterpartName", "Nome Contraparte", "Nome do fornecedor/cliente extraido", "Staging", "StagingEntry", "String", "Nao", "-", "Fornecedor ABC", "Extrato / NF", "Exibicao, vinculo manual", "Baixo", ""],
    # Settlement
    ["Settlement.amount", "Valor Liquidado", "Valor efetivamente pago/recebido nesta baixa", "Financeiro", "Settlement", "Decimal", "Sim", "-", "500.00", "Dialog de liquidacao", "Atualiza paidAmount do entry", "Critico", "Pode ser parcial"],
    ["Settlement.date", "Data Liquidacao", "Data efetiva do pagamento/recebimento", "Financeiro", "Settlement", "DateTime", "Sim", "-", "2026-03-20", "Dialog de liquidacao", "Fluxo de caixa realizado", "Alto", ""],
    # BankAccount
    ["currentBalance", "Saldo Atual", "Saldo corrente da conta bancaria", "Cadastros", "BankAccount", "Decimal", "Sim", "-", "45230.50", "Atualizado a cada Settlement", "Dashboard, fluxo de caixa", "Critico", "Atualizado atomicamente"],
    ["initialBalance", "Saldo Inicial", "Saldo no momento do cadastro", "Cadastros", "BankAccount", "Decimal", "Sim", "-", "10000.00", "Cadastro manual", "Referencia historica", "Medio", ""],
    # RecurringRule
    ["frequency", "Frequencia", "Periodicidade da geracao", "Financeiro", "RecurringRule", "Enum", "Sim", "-", "MONTHLY", "Cadastro da regra", "Calculo do proximo date", "Alto", "DAILY a ANNUAL"],
    ["nextGenerationDate", "Proximo Geracao", "Data da proxima geracao automatica", "Financeiro", "RecurringRule", "DateTime", "Sim", "-", "2026-04-01", "Calculado automaticamente", "Cron job verifica diariamente", "Alto", ""],
    # ClassificationRule
    ["priority", "Prioridade", "Ordem de avaliacao da regra (menor = primeiro)", "Config", "ClassificationRule", "Int", "Sim", "-", "10", "Cadastro manual", "Ordem de match na classificacao", "Medio", "Primeiro match ganha"],
    ["field", "Campo de Match", "Qual campo usar para comparacao", "Config", "ClassificationRule", "Enum", "Sim", "-", "CNPJ", "Cadastro manual", "Define tipo de comparacao", "Medio", "CNPJ, DESCRIPTION, VALUE_RANGE"],
    ["pattern", "Padrao de Match", "Valor ou expressao para comparar", "Config", "ClassificationRule", "String", "Sim", "-", "12345678000190 ou 'ALUGUEL' ou '1000-5000'", "Cadastro manual", "Comparado com dados do staging", "Medio", "Formato depende do field"],
    # AuditLog
    ["tableName", "Nome Tabela", "Qual tabela foi alterada", "Controles", "AuditLog", "String", "Sim", "-", "OfficialEntry", "Automatico", "Filtro no log de auditoria", "Alto", ""],
    ["action", "Acao", "Tipo de alteracao: criacao, edicao ou exclusao", "Controles", "AuditLog", "Enum", "Sim", "-", "UPDATE", "Automatico", "Filtro e rastreabilidade", "Alto", "CREATE, UPDATE, DELETE"],
    ["oldValues", "Valores Anteriores", "Estado antes da alteracao (JSON)", "Controles", "AuditLog", "Json", "Nao", "-", '{"status":"OPEN"}', "Automatico", "Comparacao de mudancas", "Alto", "Nem sempre preenchido"],
    ["newValues", "Valores Novos", "Estado apos a alteracao (JSON)", "Controles", "AuditLog", "Json", "Nao", "-", '{"status":"SETTLED"}', "Automatico", "Comparacao de mudancas", "Alto", ""],
]
add_sheet(wb, "02-Dicionario de Dados", headers, dd)

# =====================================================
# ABA 7: MODULOS E FUNCIONALIDADES
# =====================================================
headers = ["Modulo", "Funcionalidade", "Descricao", "Objetivo", "Usuario", "Frequencia", "Dependencias", "Regras", "Tabelas", "Criticidade", "Melhoria Sugerida"]
mods = [
    ["Dashboard", "KPIs Financeiros", "4 cards com receita, despesa, saldo e vencidos", "Visao rapida da saude financeira", "Todos", "Diaria", "OfficialEntry populado", "RN07, RN08", "OfficialEntry, BankAccount", "Alta", "Adicionar seletor de periodo"],
    ["Dashboard", "Graficos", "Receita vs Despesa, Fluxo de Caixa, Despesas por conta, Aging", "Analise visual de tendencias", "Todos", "Diaria", "Dados historicos", "-", "OfficialEntry", "Media", "Adicionar drill-down nos graficos"],
    ["Dashboard", "Multi-Tenant", "Visao consolidada de todas as empresas", "Gestao de multiplos clientes (BPO)", "Admin/Controller", "Diaria", "Multiplas empresas", "-", "Tenant, Membership, StagingEntry, OfficialEntry", "Alta", ""],
    ["Cadastros", "Plano de Contas", "CRUD hierarquico de contas contabeis + templates", "Estruturar classificacao contabil", "Admin", "Setup/Rara", "Nenhuma", "-", "ChartOfAccount", "Alta", ""],
    ["Cadastros", "Fornecedores", "CRUD de fornecedores com CNPJ/CPF", "Cadastrar parceiros de pagamento", "Analyst+", "Ocasional", "Nenhuma", "RN14", "Supplier", "Media", "Validar digito CNPJ/CPF"],
    ["Cadastros", "Clientes", "CRUD de clientes com CNPJ/CPF", "Cadastrar parceiros de recebimento", "Analyst+", "Ocasional", "Nenhuma", "RN14", "Customer", "Media", "Validar digito CNPJ/CPF"],
    ["Cadastros", "Contas Bancarias", "CRUD com saldo e tipo de conta", "Registrar contas para conciliacao", "Admin", "Rara", "Nenhuma", "-", "BankAccount", "Alta", ""],
    ["Cadastros", "Formas Pagamento", "CRUD com tipo, prazo e taxa", "Parametrizar meios de pagamento", "Admin", "Rara", "Nenhuma", "-", "PaymentMethod", "Baixa", "Taxa nunca e aplicada (BUG)"],
    ["Importacoes", "Extrato OFX/CSV", "Upload e parse de extratos bancarios", "Alimentar staging com dados do banco", "Analyst+", "Diaria", "Conta bancaria cadastrada", "RN10", "ImportBatch, BankStatementLine, StagingEntry", "Alta", "Adicionar deteccao de duplicatas"],
    ["Importacoes", "NF-e XML/CSV", "Upload e parse de notas fiscais", "Importar dados fiscais", "Analyst+", "Diaria", "Nenhuma", "RN10", "ImportBatch, TaxInvoiceLine, StagingEntry", "Alta", ""],
    ["Importacoes", "Pluggy Sync", "Sincronizacao automatica via Open Banking", "Automatizar importacao de extratos", "Controller+", "Diaria", "Conexao Pluggy ativa", "RN10", "PluggyConnection, ImportBatch, BankStatementLine, StagingEntry", "Alta", "Autenticar webhook"],
    ["Importacoes", "QIVE Sync", "Sincronizacao automatica de NF-e", "Automatizar importacao de notas", "Controller+", "Diaria", "Conexao QIVE ativa", "RN10", "QiveConnection, ImportBatch, TaxInvoiceLine, StagingEntry", "Media", ""],
    ["Staging", "Classificacao", "Classificar lancamentos automatica e manualmente", "Preparar lancamentos para incorporacao", "Analyst+", "Diaria", "Regras de classificacao", "RN10, RN01", "StagingEntry, ClassificationRule", "Alta", ""],
    ["Staging", "Validacao", "Verificar completude e regras", "Garantir qualidade dos dados", "Analyst+", "Diaria", "Classificacao feita", "RN01, RN02, RN03", "StagingEntry", "Alta", "Adicionar restricao de papel"],
    ["Staging", "Incorporacao", "Promover staging para oficial", "Oficializar lancamentos", "Controller+", "Diaria", "Validacao feita", "RN04, RN05, RN06", "StagingEntry, OfficialEntry", "Critica", "Corrigir race condition seq num"],
    ["Financeiro", "Lancamentos", "Visao central de todos lancamentos oficiais", "Gerenciar contas a pagar/receber", "Todos", "Diaria", "Incorporacao feita", "RN07, RN09", "OfficialEntry, Settlement", "Critica", "Corrigir chamadas API inexistentes"],
    ["Financeiro", "Liquidacao", "Registrar pagamentos/recebimentos", "Baixar titulos financeiros", "Analyst+", "Diaria", "Lancamento OPEN/PARTIAL", "RN07, RN08", "Settlement, OfficialEntry, BankAccount", "Critica", "Verificar periodo bloqueado"],
    ["Financeiro", "Parcelamento", "Dividir lancamento em parcelas", "Facilitar pagamento/recebimento parcelado", "Analyst+", "Ocasional", "Lancamento OPEN", "RN09", "OfficialEntry", "Alta", ""],
    ["Conciliacao", "Bancaria Auto", "Match automatico extrato vs liquidacoes", "Automatizar conciliacao", "Analyst+", "Diaria", "Extrato + liquidacoes", "RN11", "Reconciliation, BankStatementLine, Settlement", "Alta", "Adicionar match com OfficialEntry direto"],
    ["Conciliacao", "Bancaria Manual", "Match manual de itens nao conciliados", "Conciliar itens complexos", "Analyst+", "Diaria", "Itens nao conciliados", "-", "Reconciliation", "Media", "Validar pertencimento ao tenant"],
    ["Relatorios", "DRE", "Demonstracao do Resultado do Exercicio", "Visao de resultado contabil", "Controller+", "Mensal", "Lancamentos SETTLED no periodo", "-", "OfficialEntry, ChartOfAccount", "Alta", "Adicionar seletor de ano"],
    ["Relatorios", "Aging", "Analise de aging (vencidos por faixa)", "Controle de inadimplencia", "Controller+", "Semanal", "Lancamentos OPEN/PARTIAL", "-", "OfficialEntry", "Alta", "Adicionar drill-down"],
    ["Relatorios", "Orcado vs Realizado", "Comparacao orcamento x realizado", "Controle orcamentario", "Controller+", "Mensal", "BudgetLine e lancamentos", "-", "BudgetLine, OfficialEntry", "Media", "Adicionar edicao de budget"],
    ["Controles", "Check Diario", "5 verificacoes automaticas de saude", "Detectar problemas cedo", "Controller+", "Diaria", "Dados existentes", "-", "Varias", "Alta", "Adicionar links para itens afetados"],
    ["Controles", "Log Auditoria", "Historico de todas alteracoes", "Rastreabilidade e compliance", "Admin", "Sob demanda", "Operacoes de escrita", "-", "AuditLog", "Critica", "Adicionar diff de valores"],
    ["Controles", "Excecoes", "Deteccao de outliers e falhas", "Controle de qualidade dos dados", "Controller+", "Semanal", "Dados existentes", "-", "OfficialEntry, StagingEntry", "Alta", ""],
    ["Estoque", "Movimentacoes", "Historico de entradas/saidas de estoque", "Controle de estoque", "Analyst+", "Diaria", "Produtos e depositos", "-", "StockMovement, Product, Warehouse", "Media", "Adicionar CRUD"],
    ["Estoque", "Kardex", "Ficha de estoque por produto", "Rastreamento individual", "Analyst+", "Sob demanda", "Movimentacoes existentes", "-", "StockMovement, Product", "Media", ""],
    ["Config", "Usuarios", "Gerenciar membros e papeis", "Controle de acesso", "Admin", "Ocasional", "Nenhuma", "RN15", "Membership, TenantInvite, User", "Critica", ""],
    ["Config", "Empresas", "Gerenciar empresas multi-tenant", "Administrar clientes BPO", "Admin", "Rara", "Nenhuma", "RN16", "Tenant, Membership", "Alta", ""],
]
add_sheet(wb, "07-Modulos Funcionalidades", headers, mods)

# =====================================================
# ABA 8: TELAS / INTERFACES
# =====================================================
headers = ["Tela", "Objetivo", "Modulo", "Rota", "Usuario", "Campos Exibidos", "Acoes", "Validacoes", "Tabelas Lidas", "Tabelas Escritas", "Risco Uso Incorreto", "Melhoria"]
screens = [
    ["Dashboard Principal", "Visao geral financeira com KPIs e graficos", "Dashboard", "/dashboard", "Todos", "Receita, Despesa, Saldo, Vencidos, graficos", "Nenhuma (read-only)", "-", "OfficialEntry, BankAccount", "-", "Baixo", "Seletor de periodo"],
    ["Dashboard Multi-Tenant", "Visao consolidada de todas empresas", "Dashboard", "/dashboard/multi-tenant", "Admin/Controller", "Empresa, CNPJ, papel, pendencias, excecoes", "Navegar para empresa", "-", "Tenant, Membership, StagingEntry, OfficialEntry", "-", "Baixo", ""],
    ["Staging", "Classificar e incorporar lancamentos", "Staging", "/staging", "Analyst+", "Data, descricao, valor, tipo, contraparte, status, classificacao", "Editar, Validar, Rejeitar, Incorporar (batch)", "RN01-RN06", "StagingEntry + lookups", "StagingEntry, OfficialEntry", "Alto - incorporacao sem validacao adequada", "Busca textual, filtro data"],
    ["Lancamentos Oficiais", "Gerenciar todos lancamentos financeiros", "Financeiro", "/financial/entries", "Todos", "Seq, data, descricao, valor, tipo, categoria, status, vencimento, pago", "Liquidar, Cancelar, Parcelar", "RN07, RN09", "OfficialEntry + lookups", "OfficialEntry, Settlement, BankAccount", "Critico - cancelamento sem checagem tenant", "Paginacao, busca, export"],
    ["Contas a Pagar", "Visualizar contas a pagar", "Financeiro", "/financial/payables", "Todos", "Vencimento, descricao, fornecedor, conta, valor, status", "Nenhuma (read-only)", "-", "OfficialEntry", "-", "Baixo", "Adicionar acoes, filtros, subtotais"],
    ["Contas a Receber", "Visualizar contas a receber", "Financeiro", "/financial/receivables", "Todos", "Vencimento, descricao, cliente, conta, valor, status", "Nenhuma (read-only)", "-", "OfficialEntry", "-", "Baixo", "Adicionar acoes, filtros, subtotais"],
    ["Parcelas", "Acompanhar parcelas", "Financeiro", "/financial/installments", "Todos", "Grupo, descricao, parcelas, progresso, valores", "Nenhuma (read-only)", "-", "OfficialEntry", "-", "Baixo", "Adicionar liquidacao por parcela"],
    ["Calendario Financeiro", "Visualizar vencimentos futuros (90d)", "Financeiro", "/financial/calendar", "Todos", "Data, entradas por dia, valores a pagar/receber", "Nenhuma (read-only)", "-", "OfficialEntry", "-", "Baixo", "Calendario visual real"],
    ["Regras Recorrentes", "Visualizar regras de geracao automatica", "Financeiro", "/financial/recurring", "Todos", "Nome, valor, frequencia, proximo date, conta, status", "Nenhuma (read-only)", "-", "RecurringRule", "-", "Baixo", "Adicionar CRUD completo"],
    ["Import Extrato", "Importar extrato bancario", "Importacoes", "/imports/bank-statements", "Analyst+", "Historico de imports, upload", "Upload arquivo, selecionar conta", "Formato arquivo", "ImportBatch, BankAccount", "ImportBatch, BankStatementLine, StagingEntry", "Medio - sem dedup", "Deteccao duplicatas"],
    ["Import NF-e", "Importar notas fiscais", "Importacoes", "/imports/tax-invoices", "Analyst+", "Historico de imports, upload", "Upload arquivo", "Formato arquivo", "ImportBatch", "ImportBatch, TaxInvoiceLine, StagingEntry", "Baixo", ""],
    ["Pluggy", "Gerenciar conexoes Open Banking", "Integracoes", "/imports/pluggy", "Controller+", "Conexoes, status, ultimo sync", "Conectar, vincular conta, sync, desconectar", "Conexao ativa", "PluggyConnection, BankAccount", "PluggyConnection + sync tables", "Medio", ""],
    ["Conciliacao Bancaria", "Conciliar extrato com lancamentos", "Conciliacao", "/reconciliation/bank", "Analyst+", "Linhas extrato (esq) vs lancamentos (dir)", "Auto-match, match manual, desfazer", "RN11", "BankStatementLine, Settlement, Reconciliation", "Reconciliation", "Medio", ""],
    ["DRE", "Demonstracao de Resultado", "Relatorios", "/reports/income-statement", "Controller+", "Contas, valores mensais Jan-Dez, YTD", "Nenhuma (read-only)", "-", "OfficialEntry, ChartOfAccount", "-", "Baixo", "Seletor ano, export PDF"],
    ["Aging", "Analise de inadimplencia", "Relatorios", "/reports/aging", "Controller+", "Faixas de atraso, totais", "Nenhuma (read-only)", "-", "OfficialEntry", "-", "Baixo", "Drill-down, export"],
    ["Log Auditoria", "Historico de alteracoes", "Controles", "/controls/audit-log", "Admin", "Data/hora, usuario, tabela, acao, ID registro", "Filtrar", "-", "AuditLog", "-", "Baixo", "Mostrar diff, paginacao"],
    ["Excecoes", "Deteccao de anomalias", "Controles", "/controls/exceptions", "Controller+", "Severidade, tipo, descricao, fonte, valor", "Nenhuma (read-only)", "-", "OfficialEntry, StagingEntry", "-", "Baixo", "Links para itens"],
    ["Usuarios", "Gerenciar acesso", "Config", "/settings/users", "Admin", "Nome, email, papel, convites", "Convidar, alterar papel, remover", "RN15", "Membership, User, TenantInvite", "Membership, TenantInvite", "Alto - erro no papel da consequencias", ""],
    ["Empresas", "Gerenciar empresas", "Config", "/settings/companies", "Admin", "Empresas, CNPJ, stats, excecoes", "Criar, trocar, onboarding", "RN16", "Tenant, Membership", "Tenant, Membership", "Medio", ""],
]
add_sheet(wb, "08-Telas Interfaces", headers, screens)

# =====================================================
# ABA 9: ENTRADAS, PROCESSAMENTOS E SAIDAS
# =====================================================
headers = ["Processo", "Entrada", "Processamento", "Saida", "Gatilho", "Validacoes", "Dependencias", "Erro Possivel", "Tratamento Erro", "Impacto Operacional", "Obs"]
eps = [
    ["Importar Extrato", "Arquivo OFX/CSV", "Parse, cria BankStatementLine + StagingEntry, auto-classifica", "N lancamentos no staging", "Upload pelo usuario", "Formato arquivo, conta selecionada", "BankAccount cadastrada", "Arquivo mal formatado, colunas incorretas", "ImportBatch marcado FAILED", "Dados nao importados, retrabalho", ""],
    ["Importar NF-e", "Arquivo XML/CSV", "Parse, cria TaxInvoiceLine + StagingEntry", "N lancamentos + dados fiscais", "Upload pelo usuario", "XML valido, accessKey", "Nenhuma", "XML corrompido, NF duplicada", "Erro por NF, batch continua", "NFs nao importadas", "Dedup por accessKey"],
    ["Sync Pluggy", "API Open Banking", "Fetch transacoes, dedup, cria BSL + Staging", "Extrato automatico + staging", "Botao sync ou webhook", "Conexao ativa, conta vinculada", "PluggyConnection, BankAccount", "API indisponivel, erro de conexao", "Connection marcada ERROR", "Dados desatualizados", ""],
    ["Auto-Classificacao", "StagingEntry PENDING", "Avalia regras por prioridade, aplica primeira match", "Staging com classificacao preenchida", "Automatico pos-importacao ou manual", "Regras ativas existem", "ClassificationRule configuradas", "Nenhuma regra corresponde", "Entry permanece PENDING", "Classificacao manual necessaria", ""],
    ["Validacao Staging", "StagingEntry classificado", "7 verificacoes: campos, duplicata, periodo", "Status VALIDATED ou lista de erros", "Botao validar na tela", "RN01", "Classificacao previa", "Campo faltando, duplicata, periodo bloqueado", "Erros retornados ao usuario", "Lancamento nao avanca pipeline", ""],
    ["Incorporacao", "StagingEntry VALIDATED", "Cria OfficialEntry, num sequencial, muda status", "Lancamento oficial criado", "Botao incorporar (ADMIN/CONTROLLER)", "Status VALIDATED", "Validacao concluida", "Race condition no seq number", "Erro de unicidade no banco", "Numeros duplicados possivel", "RN04-06"],
    ["Liquidacao", "OfficialEntry OPEN/PARTIAL", "Cria Settlement, atualiza status e saldo bancario", "Titulo baixado, saldo atualizado", "Dialog na tela de lancamentos", "Valor <= restante", "Conta bancaria, lancamento aberto", "Valor excede saldo restante", "Erro retornado ao usuario", "Titulo nao baixado", "RN07-08"],
    ["Parcelamento", "OfficialEntry OPEN", "Cancela original, cria N novos", "N parcelas escalonadas", "Dialog na tela de lancamentos", "Status OPEN, N>=2", "Lancamento aberto", "Arredondamento de centavos", "Ultimo recebe diferenca", "Centavo extra na ultima parcela", "RN09"],
    ["Conciliacao Auto", "BSL + Settlement nao conciliados", "3 passes: exato, data +/-2d, valor +/-1%", "Registros de conciliacao criados", "Botao auto-reconcile", "Conta bancaria selecionada", "Extrato e liquidacoes existem", "Nenhum match encontrado", "Itens permanecem abertos", "Conciliacao manual necessaria", "RN11"],
    ["Cron Recorrentes", "RecurringRule ativas devidas", "Cria StagingEntry por regra, atualiza next date", "Novos staging entries", "Cron diario automatico", "CRON_SECRET", "Regras ativas", "Falha numa regra para todas", "Sem isolamento de erro", "Regras restantes nao processadas", "RN12"],
    ["Cron Vencidos", "OfficialEntry vencidos", "Busca entries vencidos, cria AuditLog", "Registros no log de auditoria", "Cron diario automatico", "CRON_SECRET", "Lancamentos existentes", "Duplicatas de log", "Sem dedup", "Logs repetidos", "RN13"],
]
add_sheet(wb, "09-Entradas Process Saidas", headers, eps)

# =====================================================
# ABA 10: INTEGRACOES
# =====================================================
headers = ["Integracao", "Sistema Origem", "Sistema Destino", "Tipo", "Dados Trafegados", "Frequencia", "Validacao", "Risco", "Contingencia", "Impacto Falha", "Obs"]
integs = [
    ["Pluggy Open Banking", "Bancos (via Pluggy)", "Sistema Financeiro", "API REST (pull + webhook push)", "Transacoes bancarias: data, valor, descricao, saldo", "Diaria (manual ou webhook)", "Dedup por externalId, conexao ativa", "Alto - webhook sem autenticacao, API pode ficar indisponivel", "Importacao manual de OFX/CSV como fallback", "Extratos desatualizados, staging nao alimentado", "Token de acesso via Pluggy SDK"],
    ["QIVE NFe", "SEFAZ (via QIVE)", "Sistema Financeiro", "API REST (pull)", "Notas fiscais eletronicas: chave acesso, emitente, valores, impostos", "Diaria (manual)", "Dedup por accessKey, credenciais validas", "Medio - API pode mudar, credenciais globais (nao por tenant)", "Importacao manual de XML como fallback", "NFs nao importadas automaticamente", "Cursor de paginacao persistido"],
    ["NextAuth (Credentials)", "Formulario de Login", "JWT Session", "Interno", "Email, senha hasheada, papel do usuario", "A cada login", "Email existe, senha confere (bcrypt)", "Medio - sem rate limiting, sem 2FA", "Reset de senha (futuro)", "Acesso nao autorizado", "JWT strategy, sem DB sessions"],
    ["Cron Recorrentes", "RecurringRule (banco)", "StagingEntry (banco)", "Cron Job interno", "Dados da regra -> novo staging entry", "Diaria", "CRON_SECRET no header", "Alto - sem transacao, sem isolamento de erro", "Execucao manual", "Lancamentos nao gerados", "POST /api/cron/recurring-entries"],
    ["Cron Vencidos", "OfficialEntry (banco)", "AuditLog (banco)", "Cron Job interno", "IDs de entries vencidos -> registros de log", "Diaria", "CRON_SECRET no header", "Baixo - apenas registra", "Check manual", "Vencidos nao detectados no log", "POST /api/cron/overdue-check"],
    ["Vercel Deploy", "Repositorio Git", "Vercel (hosting)", "CI/CD", "Codigo fonte, build artifacts", "A cada push", "Build precisa passar", "Baixo", "Rollback pelo Vercel", "Sistema fora do ar", "Next.js build + deploy"],
    ["Neon PostgreSQL", "Sistema Financeiro", "Banco de Dados", "Prisma ORM", "Todas as tabelas e dados", "Contínua", "Schema sync via migrations", "Medio - cold starts do Neon", "Connection pooling", "Queries lentas no cold start", "Regiao sa-east-1"],
]
add_sheet(wb, "10-Integracoes", headers, integs)

# =====================================================
# ABA 11: PERFIS E PERMISSOES
# =====================================================
headers = ["Perfil", "Modulo", "Funcionalidade", "Nivel Acesso", "Acoes Bloqueadas", "Risco", "Recomendacao", "Obs"]
perms = [
    ["ADMIN", "Todos", "Acesso total", "Total", "Nenhuma", "Alto - pode causar dano maximo", "Limitar a 1-2 usuarios por empresa", "Nivel hierarquico 4"],
    ["ADMIN", "Config", "Criar empresa, convidar usuarios, alterar papeis", "Escrita", "Nenhuma", "Alto", "Revisar convites periodicamente", ""],
    ["ADMIN", "Config", "Aplicar templates (plano de contas, produtos)", "Escrita", "Nenhuma", "Alto - pode limpar dados existentes", "Fazer backup antes", "Opcao de limpar existentes e perigosa"],
    ["ADMIN", "Integracoes", "Deletar conexao QIVE", "Escrita", "Nenhuma", "Medio", "", ""],
    ["CONTROLLER", "Staging", "Incorporar lancamentos", "Escrita", "Criar empresa, convidar usuarios, aplicar templates", "Medio", "Validar lancamentos antes de incorporar", "Nivel hierarquico 3"],
    ["CONTROLLER", "Integracoes", "Sync Pluggy, Sync QIVE, Deletar Pluggy", "Escrita", "Deletar QIVE, gerenciar usuarios", "Medio", "", ""],
    ["ANALYST", "Staging", "Classificar, validar, rejeitar", "Escrita", "Incorporar lancamentos, sync integracoes, deletar conexoes, gerenciar usuarios", "Baixo", "Pode classificar errado sem revisao", "Nivel hierarquico 2"],
    ["ANALYST", "Importacoes", "Todas importacoes de arquivo", "Escrita", "Sync automatico", "Medio", "Verificar arquivo antes de importar", ""],
    ["ANALYST", "Financeiro", "Liquidar, parcelar, cancelar lancamentos", "Escrita", "Incorporar", "Alto - BUG: deveria ser restrito", "Adicionar restricao de papel na liquidacao", "Problema de seguranca atual"],
    ["ANALYST", "Cadastros", "CRUD completo de todos cadastros", "Escrita", "Templates", "Medio", "Revisar mudancas em plano de contas", ""],
    ["VIEWER", "Todos", "Somente visualizacao", "Leitura", "Todas as escritas", "Baixo", "Papel padrao para novos usuarios", "Nivel hierarquico 1"],
    ["VIEWER", "Dashboard", "Ver KPIs e graficos", "Leitura", "Filtrar, exportar", "Baixo", "", ""],
    ["VIEWER", "Relatorios", "Ver DRE, Aging, Orcado vs Realizado", "Leitura", "Exportar", "Baixo", "", ""],
    ["VIEWER", "Financeiro", "BUG: Pode liquidar/cancelar", "Escrita (BUG)", "Deveria ser bloqueado", "CRITICO", "CORRIGIR: adicionar requireRole na liquidacao e cancelamento", "Falha de seguranca"],
    ["SISTEMA (Cron)", "Financeiro", "Gerar recorrentes, detectar vencidos", "Escrita", "Tudo exceto crons", "Medio", "Monitorar logs do cron", "Autenticado por CRON_SECRET"],
]
add_sheet(wb, "11-Perfis Permissoes", headers, perms)

# =====================================================
# ABA 12: LOGS, AUDITORIA E RASTREABILIDADE
# =====================================================
headers = ["Evento", "Origem", "Tabela Afetada", "Usuario", "Tipo Alteracao", "Info Anterior", "Info Nova", "Criticidade", "Retencao", "Cobertura Atual", "Obs"]
logs = [
    ["Criar fornecedor", "Tela de Fornecedores", "Supplier", "Quem criou", "CREATE", "-", "Dados completos (JSON)", "Media", "Indefinida", "Coberto", ""],
    ["Editar fornecedor", "Tela de Fornecedores", "Supplier", "Quem editou", "UPDATE", "Valores anteriores (JSON)", "Valores novos (JSON)", "Media", "Indefinida", "Coberto", "Grava diff"],
    ["Excluir fornecedor (soft)", "Tela de Fornecedores", "Supplier", "Quem excluiu", "DELETE", "active: true", "active: false", "Media", "Indefinida", "Coberto", "Na verdade e UPDATE de active"],
    ["Criar lancamento oficial", "Incorporacao", "OfficialEntry", "Quem incorporou", "CREATE", "-", "IDs criados", "Alta", "Indefinida", "Parcial - registra IDs nao dados completos", "Via service staging.ts"],
    ["Liquidar lancamento", "Tela de Lancamentos", "Settlement + OfficialEntry", "Quem liquidou", "CREATE + UPDATE", "-", "Settlement criado, status atualizado", "Critica", "Indefinida", "Parcial - service cria audit mas nao a action", ""],
    ["Cancelar lancamento", "Tela de Lancamentos", "OfficialEntry", "-", "UPDATE", "-", "status: CANCELLED", "Critica", "Indefinida", "NAO COBERTO (BUG)", "Sem audit log, sem rastreabilidade"],
    ["Validar staging", "Tela de Staging", "StagingEntry", "Quem validou", "UPDATE", "Status anterior", "VALIDATED", "Alta", "Indefinida", "Coberto", "Via service staging.ts"],
    ["Rejeitar staging", "Tela de Staging", "StagingEntry", "-", "UPDATE", "-", "REJECTED + motivo", "Media", "Indefinida", "NAO COBERTO (BUG)", ""],
    ["Importar arquivo", "Tela de Importacao", "ImportBatch", "Quem importou", "CREATE", "-", "Dados do batch", "Alta", "Indefinida", "Coberto", ""],
    ["Conciliacao automatica", "Tela Conciliacao", "Reconciliation", "Quem executou", "CREATE", "-", "Quantidade de matches", "Media", "Indefinida", "Coberto", "Via service reconciliation.ts"],
    ["Desfazer conciliacao", "Tela Conciliacao", "Reconciliation", "-", "DELETE", "-", "-", "Media", "Indefinida", "NAO COBERTO (BUG)", ""],
    ["Criar empresa", "Tela Empresas", "Tenant", "Quem criou", "CREATE", "-", "Dados da empresa", "Alta", "Indefinida", "Coberto", ""],
    ["Convidar usuario", "Tela Usuarios", "TenantInvite", "Quem convidou", "CREATE", "-", "Email + papel", "Alta", "Indefinida", "Coberto", ""],
    ["Trocar empresa ativa", "Header", "Membership", "Quem trocou", "UPDATE", "-", "-", "Media", "Indefinida", "NAO COBERTO", ""],
    ["Login/Logout", "Tela Login", "Session", "Quem logou", "-", "-", "-", "Critica", "Indefinida", "NAO COBERTO", "NextAuth nao grava por padrao"],
    ["Deteccao vencidos (cron)", "Cron Job", "AuditLog", "system@cron", "UPDATE", "-", "Deteccao de atraso", "Media", "Indefinida", "Coberto mas gera duplicatas", ""],
    ["Geracao recorrente (cron)", "Cron Job", "StagingEntry + RecurringRule", "Sistema", "CREATE + UPDATE", "-", "-", "Alta", "Indefinida", "NAO COBERTO (BUG)", "Service nao cria audit"],
]
add_sheet(wb, "12-Logs Auditoria", headers, logs)

wb.save(WB_PATH)
print(f"Parte 2 salva: {WB_PATH}")
print(f"Abas: {wb.sheetnames}")
